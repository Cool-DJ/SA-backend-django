import io
import json
import logging
import uuid
from datetime import datetime, timezone
import logging
from pac.rrf.service_matrix_validation import validate_new_lane, validate_new_lane
from pac.rrf.validators import do_validate_lanes, do_validate_pricing_point, init_lane_validators, resolve_lane_ids, \
    resolve_lane_ids_orm, resolve_pricing_point_ids

logging.getLogger().setLevel(logging.INFO)

import openpyxl
from dateutil.relativedelta import relativedelta
from django.db import connection, transaction
from django.db.models import Count
from django.http import JsonResponse, HttpResponse
from django_filters import rest_framework as filters
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle, Spacer
from rest_framework import generics, mixins, status, views, viewsets
from rest_framework.decorators import action, api_view
from rest_framework.filters import OrderingFilter
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from xlsxwriter import Workbook

from pac.rrf.utils import str2bool

import pac.rrf.queries as queries
from core.models import User
from pac.helpers.connections import pyodbc_connection, getFormattedRowResults
from pac.helpers.functions import revert_instance
from pac.helpers.mixins import (GetQuerySetMixin, GetSerializerClassMixin,
                                RetrieveHistoryMixin,
                                RevertVersionMixin)
from pac.models import Account, BasingPoint, Country, LocationTreeView, Notification, PostalCode, Province, Region, \
    ServicePoint, \
    Terminal, WeightBreakHeader, ServiceLevel
from pac.notifications import NotificationManager
from pac.rrf import enums
from pac.rrf.lanes.request_section_lane import RequestSectionLaneView
from pac.rrf.tasks import validate_lanes, validate
from pac.rrf.models import (Customer, CustomerZones, PointType, Request,
                            RequestEditorRight, RequestHistory,
                            RequestInformation, RequestProfile,
                            RequestProfileHistory, Review,
                            RequestSection, RequestSectionHistory,
                            RequestSectionLane, RequestSectionLaneHistory, RequestSectionLanePointType,
                            RequestStatusType, RequestSectionLaneImportQueue, ImportFile,
                            RequestSectionLanePricingPoint, RequestSectionLanePricingPointImportQueue, Language)
from pac.rrf.serializers import (AccountOwnerSerializer, CustomerSerializer,
                                 RequestHistorySerializer,
                                 RequestInformationSerializer,
                                 RequestProfileHistoryRetrieveSerializer,
                                 RequestProfileRetrieveSerializer,
                                 RequestProfileSerializer,
                                 RequestSectionHistoryRetrieveSerializer,
                                 RequestSectionLaneHistoryRetrieveSerializer, RequestSectionLaneImportQueueSerializer,
                                 RequestSectionLaneSerializer,
                                 RequestSectionRetrieveSerializer,
                                 RequestSectionSerializer, RequestSerializer,
                                 RequestSectionLanePricingPointSerializer)
from pac.rrf.utils import request_section_lane_insert, request_section_lane_pricingpiont_insert, all_elem_same_value
from pac.rrf.workflow.workflow import WorkflowManager, request_editor_right_update
from pac.utils import save_comment
from django.conf.urls import include, url
from pac.rrf.urls import path
import pac.validation as validation
from pac.rrf.review.review import ReviewAPI

# import settings for key vault variables
from django.conf import settings


class CreateRequestTariff(views.APIView):

    def post(self, request):
        try:
            user = self.request.user
            data = request.data
            request_type = data.get("RequestType")
            service_level_id = data.get("ServiceLevelID")
            account_id = data.get("AccountID")
            conn = pyodbc_connection()
            cursor = conn.cursor()
            lang_name = 'English'
            print(f'creating request: {service_level_id} {request_type} user {user.user_id}')
            if request_type == 'non_customer':
                dummy_account_id = Account.objects.filter(account_number=0).first()
                uni_type = "Non-Customer Tariff"
                cursor.execute("EXEC [dbo].[Request_By_Account_Select] ?, ?, ?, ?, ?, ?",
                               user.user_id, service_level_id, dummy_account_id.account_id, None, uni_type, lang_name)
            elif request_type == 'customer':
                uni_type = "Customer Tariff"
                cursor.execute("EXEC [dbo].[Request_By_Account_Select] ?, ?, ?, ?, ?, ?",
                               user.user_id, service_level_id, account_id, None, uni_type, lang_name)
            elif request_type == 'tender':
                uni_type = "Tender Tariff"
                cursor.execute("EXEC [dbo].[Request_By_Account_Select] ?, ?, ?, ?, ?, ?",
                               user.user_id, service_level_id, account_id, None, uni_type, lang_name)
            else:
                uni_type = "Annual Review Tariff"
                cursor.execute("EXEC [dbo].[Request_By_Account_Select] ?, ?, ?, ?, ?, ?",
                               user.user_id, service_level_id, account_id, None, uni_type, lang_name)
            raw_data = cursor.fetchone()
            cursor.commit()
            payload = json.loads(raw_data[0]) if raw_data[0] else {}
            if uni_type == "Annual Review Tariff":
                insert_review = {'RequestID': payload['request_id']}
                insert_review = [{"data": insert_review}]
                review_api = ReviewAPI()
                review_api.conn = conn
                review = review_api.bulk_insert(insert_review, kwargs=payload['request_id'])
                expiry_date = f"""
                            UPDATE dbo.Review SET ReviewExpDate = CAST(DATEADD(year, 1, GETDATE()) as date)
                            WHERE RequestID = {payload['request_id']};
                            EXEC dbo.[Audit_Record] @TableName = 'Review', @PrimaryKeyValue = {review[0]}, @UpdatedBy = '{user.user_id}';
                        """
                cursor.execute(expiry_date)
                cursor.commit()
            return Response({"status": "Success",
                             "request_id": payload['request_id']}, status=status.HTTP_200_OK)
        except Exception as e:
            logging.warning("{} {}".format(type(e).__name__, e.args))
            return Response({"status": "Failure", "error": "{} {}".format(type(e).__name__, e.args)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class rate_base_search_by_description_pyodbc(views.APIView):

    def get(self, request, *args, **kwargs):
        rate_base_description = kwargs.get("rate_base_description")
        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        query = queries.RATE_BASE_SEARCH_BY_DESCRIPTION.format(rate_base_description)

        cursor.execute(query)
        raw_data = cursor.fetchone()
        payload = json.loads(raw_data[0]) if raw_data[0] else []

        return Response(payload, status=status.HTTP_200_OK)


class SpeedsheetPyodbcView(views.APIView):

    def post(self, request, *args, **kwargs):
        try:
            user = self.request.user
            account_id = kwargs.get("account_id")
            # service_level_id = kwargs.get("service_level_id")
            service_level_id = request.data.get("service_level_id", None)
            language_code = request.data.get("language", None)
            # language = Language.objects.filter(language_code == lang_code).get()
            # request_section = RequestSection.objects.filter(request_section_id=lang_code).values_list(
            #     "weight_break", "section_name", flat=False).get()

            speedsheet_name = request.data.get("name", None)

            cnxn = pyodbc_connection()
            cursor = cnxn.cursor()
            cursor.execute("EXEC [dbo].[Request_By_Account_Select] ?, ?, ?, ?, ?, ?",
                           user.user_id, service_level_id, None, speedsheet_name, 'SPEEDSHEET', language_code)

            raw_data = cursor.fetchone()
            cursor.commit()
            payload = json.loads(raw_data[0]) if raw_data[0] else {}
            return Response(payload, status=status.HTTP_200_OK)
        except Exception as e:
            logging.warning("{} {}".format(type(e).__name__, e.args))
            return Response({"status": "Failure", "error": "{} {}".format(type(e).__name__, e.args)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get(self, request, *args, **kwargs):
        user_id = self.request.user.user_id
        # user_id = 57
        request_id = kwargs.get("request_id")
        # user_id = kwargs.get("user_id")
        # user_id = kwargs.get("user_id")
        # user_id = user_id if user_id > 0 else self.request.user.user_id

        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        if request_id is not None:
            query = queries.GET_REQUEST.format(user_id, 'SPEEDSHEET', request_id)
        else:
            query = queries.GET_SPEEDSHEET_LIST.format(user_id, 'SPEEDSHEET')

        cursor.execute(query)
        raw_data = cursor.fetchone()
        payload = json.loads(raw_data[0]) if raw_data[0] else []

        return Response(payload, status=status.HTTP_200_OK)


class NumberedCanvas(canvas.Canvas):

    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self.Canvas = canvas.Canvas
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        """add page info to each page (page x of y)"""
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            # pdb.set_trace()
            self.setFont("Helvetica", 8)
            self.draw_page_number(num_pages)
            self.Canvas.showPage(self)
        self.Canvas.save(self)

    def draw_page_number(self, page_count):
        # Change the position of this to wherever you want the page number to be
        # self.drawRightString(211 * mm, 15 * mm + (0.2 * inch),
        #                      "Page %d of %d" % (self._pageNumber, page_count))
        # self.drawCentredString(211 * mm, 15 * mm + (0.2 * inch),
        #                      "Page %d of %d" % (self._pageNumber, page_count))

        self.drawRightString(30 * mm, 10 * mm, "Page %d of %d" % (self._pageNumber, page_count))


class SpeedsheetExportPyodbcView(views.APIView):
    @staticmethod
    def _header_footer(canvas, doc):
        # Save the state of our canvas so we can draw on it
        canvas.saveState()
        styles = getSampleStyleSheet()

        # Header
        header = Paragraph('This is a multi-line header.  It goes on every page.   ' * 5, styles['Normal'])
        w, h = header.wrap(doc.width, doc.topMargin)
        header.drawOn(canvas, doc.leftMargin, doc.height + doc.topMargin - h)

        # Footer
        footer = Paragraph('This is a multi-line footer.  It goes on every page.   ' * 5, styles['Normal'])
        w, h = footer.wrap(doc.width, doc.bottomMargin)
        footer.drawOn(canvas, doc.leftMargin, h)

        # Release the canvas
        canvas.restoreState()

    def create_pdf(self, service_level_name, col_titles, lanes, initiated_on, language_code):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=' + uuid.uuid4().__str__() + ".pdf"

        story = []
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, topMargin=10, leftMargin=30, rightMargin=10)

        styles = getSampleStyleSheet()

        title = "PROPOSAL ONLY: TARIFF - SPEED SHEET"
        title_style = styles['Heading2']
        title_style.alignment = 1
        title = Paragraph(title, title_style)
        story.append(title)
        story.append(Spacer(20, 20))
        expiry_date = initiated_on + relativedelta(months=+6)

        table_data = []
        header_data = ['SERVICE LEVEL: {}'.format(service_level_name),
                       'ISSUE DATE: {}'.format(initiated_on.strftime('%m/%d/%Y'))]
        header_data = tuple(header_data)
        table_data.append(header_data)
        header_data = ['DENSITY: {}'.format("###"),
                       'EXPIRY DATE: {}'.format(expiry_date.strftime('%m/%d/%Y'))]
        header_data = tuple(header_data)
        table_data.append(header_data)

        table = Table(table_data, hAlign='LEFT', colWidths=[220, 220])
        table_style = [
            # ('GRID', (0, 0), (-1, -1), 0.25, colors.red),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (7, 0), 5),
            ('TOPPADDING', (0, 0), (7, 0), 5),
            ('FONTSIZE', (0, 0), (-1, -1), 9)
        ]
        table.setStyle(TableStyle(table_style))

        story.append(table)
        # ---

        story.append(Spacer(10, 10))

        # hr = HRFlowable(width="100%", thickness=1, lineCap='round', color=black, spaceBefore=1, spaceAfter=1,
        #                 hAlign='CENTER', vAlign='BOTTOM', dash=True)
        # story.append(hr)
        table_data = []
        header_data = ['ORIGIN', 'DESTINATION']
        header_data = tuple(header_data + col_titles)
        table_data.append(header_data)
        # story.append(hr)
        for row_num, lane in enumerate(lanes):
            # origin_row = ''
            # if row_num == 0:
            #     origin_row = lane.origin_code + ',' + lane.origin_province_code
            origin_row = lane.origin_code + ',' + lane.origin_province_code
            row_data = [origin_row,
                        lane.destination_code + ',' + lane.destination_province_code]
            for rate in get_rates(lane, 'commitment'):
                row_data.append(rate)
            row_data = tuple(row_data)
            table_data.append(row_data)
        table = Table(table_data, hAlign='LEFT', repeatRows=1)
        table_style = [

            ('LINEABOVE', (0, 0), (7, 0), 2, colors.black),
            ('LINEBELOW', (0, 0), (7, 0), 2, colors.black),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (7, 0), 30),
            ('TOPPADDING', (0, 0), (7, 0), 5),
            ('FONTSIZE', (0, 0), (-1, -1), 9)
        ]
        table.setStyle(TableStyle(table_style))

        story.append(table)

        bottom_note = Paragraph('Note: These rates will not be applied to a freight bill.', styles['Normal'])

        story.append(bottom_note)

        doc.build(story, canvasmaker=NumberedCanvas)
        # doc.build(story, onFirstPage=self._header_footer, onLaterPages=self._header_footer)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    def create_xls(self, service_level_name, col_titles, lanes):
        output_file = io.BytesIO()
        row_num = 0
        workbook = Workbook(output_file, {'in_memory': True})
        worksheet = workbook.add_worksheet(name='Pricing Points Template')

        worksheet.write(row_num, 4, 'PROPOSAL ONLY: TARIFF - SPEED SHEET')
        row_num = row_num + 3
        worksheet.write(row_num, 0, 'SERVICE LEVEL: {} DENSITY: {} ISSUE DATE: {} EXPIRY DATE: {}'.format(
            service_level_name,
            '900',
            '2021/02/07',
            '2022/02/06'
        )
                        )

        header_data = ['ORIGIN', 'DESTINATION'] + col_titles
        row_num = row_num + 3
        for col_num, col_data in enumerate(header_data):
            worksheet.write(row_num, col_num, col_data)

        for lane in lanes:
            col_num = 0
            row_num = row_num + 1
            origin_name = lane.origin_code + ',' + lane.origin_province_code
            destination_name = lane.destination_code + ',' + lane.destination_province_code

            worksheet.write(row_num, col_num, origin_name)
            col_num = col_num + 1
            worksheet.write(row_num, col_num, destination_name)

            for rate in get_rates(lane, 'commitment'):
                col_num = col_num + 1
                worksheet.write(row_num, col_num, rate)

        worksheet.write(row_num + 1, 0, 'Note: These rates will not be applied to a freight bill.')

        workbook.close()

        output_file.seek(0)

        response = HttpResponse(output_file.read(),
                                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=' + uuid.uuid4().__str__() + ".xlsx"

        output_file.close()

        return response

    def get(self, request, *args, **kwargs):

        export_format = kwargs.get("export_format")
        speed_sheet_id = kwargs.get("speed_sheet_id")
        speed_sheet = Request.objects.filter(request_id=speed_sheet_id).first()

        request_section = RequestSection.objects.filter(request_id=request_id).values_list(
            "weight_break", "section_name", "request_section_id", "weight_break_header", flat=False).get()
        initiated_on = speed_sheet.initiated_on
        language_code = speed_sheet.language_id

        service_level = WeightBreakHeader.objects.filter(weight_break_header_id=request_section[3]).values_list(
            "service_level__service_level_name", flat=False).get()
        col_titles = []
        weight_break = request_section[0]
        for weight_break in json.loads(weight_break):
            col_titles.append([x.split(':') for x in weight_break.split(' ')][0][1])

        lanes = list(RequestSectionLane.objects.filter(request_section_id=request_section[2], is_active=True))

        service_level_name = service_level[0]
        if export_format == 'pdf':
            return self.create_pdf(service_level_name, col_titles, lanes, initiated_on, language_code)
        elif export_format == 'xls':
            return self.create_xls(service_level_name, col_titles, lanes)


class SpeedsheetModifyPyodbcView(views.APIView):

    def post(self, request, *args, **kwargs):
        try:
            user = self.request.user
            speed_sheet_id = kwargs.get("speed_sheet_id")
            action_type = kwargs.get("action_type")
            if action_type == 'CANCEL':
                Request.objects.filter(request_number=speed_sheet_id).update(
                    is_active=0)
            # cnxn = pyodbc_connection()
            # cursor = cnxn.cursor()
            # cursor.execute("EXEC [dbo].[Request_By_Account_Select] ?, ?, ?, ?, ?, ?",
            #                user.user_id, service_level_id, None, speedsheet_name, 'SPEEDSHEET', language)
            #
            # raw_data = cursor.fetchone()
            # cursor.commit()
            # payload = json.loads(raw_data[0]) if raw_data[0] else {}
            return Response(status=status.HTTP_200_OK)
        except Exception as e:
            logging.warning("{} {}".format(type(e).__name__, e.args))
            return Response({"status": "Failure", "error": "{} {}".format(type(e).__name__, e.args)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GetRequestInformationByIDPyodbcView(views.APIView):

    def get(self, request, *args, **kwargs):
        try:
            request_number = kwargs.get("request_number")

            cnxn = pyodbc_connection()
            cursor = cnxn.cursor()

            cursor.execute(
                "EXEC [dbo].[Request_By_ID_Select] ?, ?", request_number, 0)

            raw_data = cursor.fetchone()
            payload = json.loads(raw_data[0]) if raw_data[0] else []

            return Response(payload, status=status.HTTP_200_OK)
        except Exception as e:
            logging.warning("{} {}".format(type(e).__name__, e.args))
            return Response({"status": "Failure", "error": "{} {}".format(type(e).__name__, e.args)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GetAccountHistoryPyodbcView(views.APIView):

    def get(self, request, *args, **kwargs):
        request_id = kwargs.get("request_id")
        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        query = queries.GET_ACCOUNT_HISTORY.format(request_id)

        cursor.execute(query)
        raw_data = cursor.fetchone()
        payload = json.loads(raw_data[0]) if raw_data[0] else []

        return Response(payload, status=status.HTTP_200_OK)


class GetTariffHistoryPyodbcView(views.APIView):

    def get(self, request, *args, **kwargs):
        request_id = kwargs.get("request_id")
        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        query = queries.GET_TARIFF_HISTORY.format(request_id)

        cursor.execute(query)
        raw_data = cursor.fetchone()
        payload = json.loads(raw_data[0]) if raw_data[0] else []

        return Response(payload, status=status.HTTP_200_OK)


class GetRequestLanesPyodbcView(views.APIView):

    def get(self, request, *args, **kwargs):
        request_id = kwargs.get("RequestID")
        lanes_query = """
            SELECT rs.RequestSectionID,rs.SectionName
                ,rsl.Cost
                ,olt.Code OriginCode
                ,dlt.Code DestinationCode
                , rsl.RequestSectionLaneID                                                      
            FROM dbo.RequestSection rs
            INNER JOIN dbo.RequestSectionLane rsl on rs.RequestSectionID = rsl.RequestSectionID
            INNER JOIN dbo.V_LocationTree olt ON olt.ID = rsl.OriginID AND olt.PointTypeID = rsl.OriginTypeID
            INNER JOIN dbo.V_LocationTree dlt ON dlt.ID = rsl.DestinationID AND dlt.PointTypeID = rsl.DestinationTypeID
            WHERE rs.RequestID = {request_id}  AND rs.IsInactiveViewable = 1
            ORDER BY rs.RequestSectionID, rsl.RequestSectionLaneID
        """.format(request_id=request_id)
        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        cursor.execute(lanes_query)
        columns = [column[0] for column in cursor.description]
        lanes = [dict(zip(columns, row)) for row in cursor.fetchall()]

        payload = lanes

        return Response(payload, status=status.HTTP_200_OK)


class GetRequestLaneLocationTreePyodbcView(views.APIView):

    def get(self, request, *args, **kwargs):
        request_section_id = kwargs.get("request_section_id")
        orig_point_type_name = kwargs.get("orig_type")
        orig_point_id = kwargs.get("orig_id")
        dest_point_type_name = kwargs.get("dest_type")
        dest_point_id = kwargs.get("dest_id")
        lane_status_name = kwargs.get("lane_status")
        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        query = queries.GET_REQUEST_LANE_LOCATION_TREE.format(
            request_section_id, orig_point_type_name, orig_point_id, dest_point_type_name, dest_point_id,
            lane_status_name, '')

        cursor.execute(query)
        raw_data = cursor.fetchone()
        payload = json.loads(raw_data[0]) if raw_data[0] else []

        return Response(payload, status=status.HTTP_200_OK)


# pac-1038
class RequestLaneExportView(views.APIView):

    def get(self, request, *args, **kwargs):
        request_section_id = kwargs.get("request_section_id")
        template = kwargs.get("template")
        request_section = RequestSection.objects.filter(request_section_id=request_section_id).values_list(
            "weight_break", "section_name", "request_section_id", flat=False).get()
        weight_break = request_section[0]
        section_name = request_section[1]

        lanes_query = RequestSectionLaneView.GET_FILTERED_QUERY.format(section_clause = f'rs.RequestSectionID = {request_section_id}')
        lanes_query = lanes_query.format(closing_clause = " ",opening_clause = " ",page_clause=" ",sort_clause=" ",where_clauses=" ")

        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        cursor.execute(lanes_query)
        raw_data = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        active_lanes = []
        for row in raw_data:
            active_lanes.append(dict(zip(columns, row)))
        lanes = RequestSectionLane.objects.filter(request_section_lane_id__in=[d['RequestSectionLaneID'] for d in active_lanes])

        output_file = io.BytesIO()
        row_num = 0
        col_num = 0
        workbook = Workbook(output_file, {'in_memory': True})
        worksheet = workbook.add_worksheet(name='Lanes Template')

        col_titles = [
            'RequestSectionID', 'SectionName', 'RequestSectionLaneID', 'OriginGroupTypeName', 'OriginGroupCode',
            'OriginPointTypeName', 'OriginPointCode', 'DestinationGroupTypeName', 'DestinationGroupCode',
            'DestinationPointTypeName', 'DestinationPointCode', 'IsBetween']

        for weight_break in json.loads(weight_break):
            col_titles.append(weight_break['LevelLowerBound'])

        for col_title in col_titles:
            worksheet.write(row_num, col_num, col_title)
            col_num += 1
        # Commitment / DrRate
        if len(lanes) > 0:
            for lane in lanes:
                row_num += 1

                origin_point_type_name = lane.origin_type_id.point_type_name
                destination_point_type_name = lane.destination_type_id.point_type_name
                origin_group_code = get_group_code(origin_point_type_name, lane.origin_id)
                destination_group_code = get_group_code(destination_point_type_name, lane.destination_id)
                origin_point_code = get_point_code(origin_point_type_name, lane.origin_id)
                destination_point_code = get_point_code(destination_point_type_name, lane.destination_id)

                worksheet.write(row_num, 0, request_section_id)
                worksheet.write(row_num, 1, section_name)
                worksheet.write(row_num, 2, lane.request_section_lane_id)
                worksheet.write(row_num, 3, get_group_type_name(origin_point_type_name))
                worksheet.write(row_num, 4, origin_group_code)
                worksheet.write(row_num, 5, origin_point_type_name)
                worksheet.write(row_num, 6, origin_point_code)  # TODO: FIX
                worksheet.write(row_num, 7, get_group_type_name(destination_point_type_name))
                worksheet.write(row_num, 8, destination_group_code)
                worksheet.write(row_num, 9, destination_point_type_name)
                worksheet.write(row_num, 10, destination_point_code)  # TODO: Fix
                worksheet.write(row_num, 11, lane.is_between)
                col_num = 12
                for rate in get_rates(lane, template):
                    worksheet.write(row_num, col_num, rate)
                    col_num += 1
        else:
            row_num += 1
            worksheet.write(row_num, 0, request_section[2])
            worksheet.write(row_num, 1, section_name)

        workbook.close()

        output_file.seek(0)

        response = HttpResponse(output_file.read(),
                                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = "attachment; filename=lanes_export.xlsx"

        output_file.close()

        return response


class UserByPartnerCarrierView(views.APIView):
    def get(self, request):
        filtered_users = list(User.objects.filter(persona__persona_name="Partner Carrier").values())
        return Response(filtered_users, status=status.HTTP_200_OK)


# pac-2263/2264
class SpeedsheetExportView(views.APIView):

    def get(self, request, *args, **kwargs):
        user_id = self.request.user.user_id
        # user_id = 57
        request_id = kwargs.get("request_id")
        # user_id = kwargs.get("user_id")
        # user_id = kwargs.get("user_id")
        # user_id = user_id if user_id > 0 else self.request.user.user_id

        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        if request_id is not None:
            query = queries.GET_REQUEST.format(user_id, 'SPEEDSHEET', request_id)
        else:
            query = queries.GET_SPEEDSHEET_LIST.format(user_id, 'SPEEDSHEET')

        cursor.execute(query)
        raw_data = cursor.fetchone()
        payload = json.loads(raw_data[0]) if raw_data[0] else []

        return Response(payload, status=status.HTTP_200_OK)


class RequestLanePricingPointExportView(views.APIView):

    def get(self, request, *args, **kwargs):
        request_section_id = kwargs.get("request_section_id")
        template = kwargs.get("template")
        request_section = RequestSection.objects.filter(request_section_id=request_section_id).values_list(
            "weight_break", "section_name", flat=False).get()
        weight_breaks = json.loads(request_section[0])

        section_name = request_section[1]
        lanes = list(RequestSectionLane.objects.filter(request_section_id=request_section_id).filter(is_active=True))
        output_file = io.BytesIO()
        row_num = 0
        col_num = 0
        workbook = Workbook(output_file, {'in_memory': True})
        worksheet = workbook.add_worksheet(name='Pricing Points Template')

        col_titles = [
            'RequestSectionID', 'SectionName', 'RequestSectionLaneID', 'OriginPointCode', 'DestinationPointCode',
            'RequestSectionLanePricingPointID', 'OriginPostCodeID', 'OriginPostalCodeName', 'DestinationPostCodeID',
            'DestinationPostalCodeName']

        for weight_break in weight_breaks:
            col_titles.append(weight_break['LevelLowerBound'])

        for col_title in col_titles:
            worksheet.write(row_num, col_num, col_title)
            col_num += 1

        row_num += 1

        for lane in lanes:
            worksheet.write(row_num, 0, request_section_id)
            worksheet.write(row_num, 1, section_name)
            worksheet.write(row_num, 2, lane.request_section_lane_id)

            origin_point_type_name = lane.origin_type_id.point_type_name
            destination_point_type_name = lane.destination_type_id.point_type_name
            origin_point_code = get_point_code(origin_point_type_name, lane.origin_id)
            destination_point_code = get_point_code(destination_point_type_name, lane.destination_id)

            worksheet.write(row_num, 3, origin_point_code)
            worksheet.write(row_num, 4, destination_point_code)

            pricing_points = RequestSectionLanePricingPoint.objects \
                .filter(request_section_lane_id=lane.request_section_lane_id).filter(is_active=True)

            for pricing_point in pricing_points:
                # row_num += 1
                # print(pricing_point)
                worksheet.write(row_num, 0, request_section_id)
                worksheet.write(row_num, 1, section_name)
                worksheet.write(row_num, 2, lane.request_section_lane_id)
                worksheet.write(row_num, 3, origin_point_code)
                worksheet.write(row_num, 4, destination_point_code)
                worksheet.write(row_num, 5, pricing_point.request_section_lane_pricing_point_id)
                worksheet.write(row_num, 6, pricing_point.origin_postal_code_id)
                worksheet.write(row_num, 7, pricing_point.origin_postal_code_name)
                worksheet.write(row_num, 8, pricing_point.destination_postal_code_id)
                worksheet.write(row_num, 9, pricing_point.destination_postal_code_name)
                col_num = 10
                rates = get_rates(pricing_point, template)
                if len(rates) == 0:
                    rates = [0] * len(weight_breaks)
                for rate in rates:
                    worksheet.write(row_num, col_num, rate)
                    col_num += 1
                row_num += 1
        workbook.close()

        output_file.seek(0)

        response = HttpResponse(output_file.read(),
                                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = "attachment; filename=pricing_points_export.xlsx"

        output_file.close()

        return response


def get_rates(lane, request_type):
    rate_values = []
    rates = {}

    # for Class:  commitment, customerDiscount, partnerDiscount
    # for Density: commitment, customerRate, drRate, partnerRate

    if request_type == 'commitment':
        rates = json.loads(lane.commitment)
    elif request_type == 'dr-rate':
        rates = json.loads(lane.dr_rate)
    elif request_type == 'customer-rate':
        rates = json.loads(lane.customer_rate)
    elif request_type == 'partner-rate':
        rates = json.loads(lane.partner_rate)
    elif request_type == 'customer-discount':
        rates = json.loads(lane.customer_discount)
    elif request_type == 'partner-discount':
        rates = json.loads(lane.partner_discount)

    for key, value in rates.items():
        rate_values.append(value)

    return rate_values


def get_group_type_name(point_type_name):
    province_points = ['Terminal', 'Basing Point', 'Service Point', 'Postal Code']
    if point_type_name in province_points:
        return 'Province'
    else:
        return 'Country'


def get_country_code(point_type_name, id):
    if point_type_name == "Country":
        return Country.objects.get(pk=id).country_code
    elif point_type_name == "Region":
        return Region.objects.get(pk=id).country.country_code
    elif point_type_name == 'Province':
        return Province.objects.get(pk=id).region.country.country_code
    elif point_type_name == "Customer Zone":
        return CustomerZones.objects.get(pk=id).service_point.province.region.country_code


def get_province_code(point_type_name, id):
    if point_type_name == "Terminal":
        return Terminal.objects.get(pk=id).basing_point_id.province_id.province_code
    elif point_type_name == "Basing Point":
        return BasingPoint.objects.get(pk=id).province_id.province_code
    elif point_type_name == "Service Point":
        return ServicePoint.objects.get(pk=id).basing_point_id.province_id.province_code
    elif point_type_name == "Postal Code":
        return PostalCode.objects.get(pk=id).basing_point_id.province_id.province_code


def get_point_code(point_type_name, id):
    if point_type_name == "Terminal":
        return Terminal.objects.get(pk=id).terminal_code
    elif point_type_name == "Basing Point":
        return BasingPoint.objects.get(pk=id).basing_point_name
    elif point_type_name == "Service Point":
        return ServicePoint.objects.get(pk=id).service_point_name
    elif point_type_name == "Postal Code":
        return PostalCode.objects.get(pk=id).postal_code_name
    elif point_type_name == "Country":
        return Country.objects.get(pk=id).country_code
    elif point_type_name == "Region":
        return Region.objects.get(pk=id).region_code
    elif point_type_name == 'Province':
        return Province.objects.get(pk=id).province_code
    elif point_type_name == "Customer Zone":
        return CustomerZones.objects.get(pk=id).customer_zone_name


def get_group_code(point_type_name, id):
    province_points = ['Terminal', 'Basing Point', 'Service Point', 'Postal Code']
    if point_type_name in province_points:
        return get_province_code(point_type_name, id)
    else:
        return get_country_code(point_type_name, id)


class RequestLanePreProcessingView(views.APIView):
    def post(self, request, *args, **kwargs):

        # Extract variables from request
        user = self.request.user
        # request_section_id = kwargs.get("request_section_id")
        # request_section = RequestSection.objects.get(pk=request_section_id)
        # sub_service_level_id = request_section.sub_service_level.sub_service_level_id
        rate_type = kwargs.get("rate_type")
        excel_file = request.FILES["import_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Lanes Template"]
        excel_data = list()

        # Iterate through Excel and populate excel data with rows
        weight_breaks_last_column = 100
        for row in worksheet.iter_rows():
            row_data = list()
            for idx, cell in enumerate(row[:weight_breaks_last_column]):
                cell_value = cell.value  # if cell.value is not None else None
                if cell_value != enums.service_column_name:  # EK: Should be 'Or None if no Status column from clean file'
                    row_data.append(cell_value)
                else:
                    weight_breaks_last_column = idx
                    break

            excel_data.append(row_data)

        # Collecting and creating arguments for old validation methods
        header = excel_data[0]
        weight_breaks = excel_data[0][12:]
        for idx, row in enumerate(excel_data):
            rates = row[12:]
            for r_idx, rate in enumerate(rates):
                try:
                    if rate:
                        rates[r_idx] = float(rate)
                    else:
                        rates[r_idx] = float(0)
                except ValueError:  # EK: This should be validated, or -1 inserted to trigger validation in async job
                    pass
        weight_break_json_array = dict(zip(weight_breaks, rates))

        lanes = []

        location_hierarchy = dict(
            RequestSectionLanePointType.objects
                .order_by('request_section_lane_point_type_name')
                .values_list('request_section_lane_point_type_name', 'location_hierarchy')
                .distinct())

        # Create file object
        file = ImportFile(file_name=excel_file.name, uni_type='LANE', request_section_id=0,
                          record_count=len(excel_data) - 1,
                          created_by=user, rate_type=rate_type)

        # Create lane import objects
        for i in range(1, len(excel_data)):
            row = excel_data[i]
            lane = RequestSectionLaneImportQueue(request_section_id=row[0],
                                                 section_name=row[1],
                                                 request_section_lane_id=row[2],
                                                 origin_group_type_name=row[3],
                                                 origin_group_code=row[4],
                                                 origin_point_type_name=row[5],
                                                 origin_point_code=row[6],
                                                 destination_group_type_name=row[7],
                                                 destination_group_code=row[8],
                                                 destination_point_type_name=row[9],
                                                 destination_point_code=row[10],
                                                 is_between=row[11],
                                                 weight_break=json.dumps(weight_break_json_array),
                                                 created_by=user,
                                                 uni_type='DATA',
                                                 file=file)
            lanes.append(lane)

        # Validate lanes
        for lane in lanes:
            request_section_id = lane.request_section_id
            resolve_lane_ids_orm(lane)
            existing_request_section_lanes = list(
                RequestSectionLane.objects.values('is_between')
                    .filter(request_section_id=request_section_id))
            if do_validate_lanes(file, header, lane, location_hierarchy, lanes=lanes,
                                 existing_request_section_lanes=existing_request_section_lanes) == 'VALID':
                # Validate against service matrix
                request_section = RequestSection.objects.get(pk=request_section_id)
                sub_service_level_id = request_section.sub_service_level.sub_service_level_id
                service_matrix_validation_results = validate_new_lane(service_level_id=sub_service_level_id,
                                                                      orig_point_type=lane.origin_point_type_name,
                                                                      origin_point_value=lane.origin_point_id,
                                                                      dest_point_type=lane.destination_point_type_name,
                                                                      destination_point_value=lane.destination_point_id)

                if service_matrix_validation_results["UNSERVICEABLE"]:
                    lane.uni_status = "UNSERVICEABLE"
                if service_matrix_validation_results["FLAGGED"]:
                    lane.uni_status = "FLAGGED"

        # Construct Validation Status Response
        response = {
            "filename": file.file_name,
            "lanes": []
        }
        for lane in lanes:
            lane_status = {
                "row_number": lane.initial_rec_order,
                "status_message": lane.status_message,
                "uni_status": lane.uni_status,
                "request_section_id": lane.request_section_id,
                "origin_group_type_id": lane.orig_group_type_id,
                "origin_group_id": lane.origin_group_id,
                "origin_group_name": lane.origin_group_code,
                "origin_point_type_id": lane.origin_point_type_id,
                "origin_point_id": lane.origin_point_id,
                "origin_point_name": lane.origin_point_code,
                "destination_group_type_id": lane.destination_group_type_id,
                "destination_group_id": lane.destination_group_id,
                "destination_group_name": lane.destination_group_code,
                "destination_point_type_id": lane.destination_point_type_id,
                "destination_point_id": lane.destination_point_id,
                "destination_point_name": lane.destination_point_code,
                "is_between": lane.is_between
            }
            response['lanes'].append(lane_status)

        return Response(data=response, status=status.HTTP_200_OK)


# pac-1841
class RequestLaneImportView(views.APIView):

    def post(self, request, *args, **kwargs):
        user = self.request.user
        request_section_id = kwargs.get("request_section_id")
        rate_type = kwargs.get("rate_type")
        excel_file = request.FILES["import_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Lanes Template"]
        # print(worksheet)
        excel_data = list()

        # 100 just an arbitrary value based on assumption that we won't have more than 100 columns.
        weight_breaks_last_column = 100
        for row in worksheet.iter_rows():
            row_data = list()
            for idx, cell in enumerate(row[:weight_breaks_last_column]):
                cell_value = cell.value  # if cell.value is not None else None
                if cell_value != enums.service_column_name:  # EK: Should be 'Or None if no Status column from clean file'
                    row_data.append(cell_value)
                else:
                    weight_breaks_last_column = idx
                    break

            excel_data.append(row_data)

        file = ImportFile(file_name=excel_file.name, uni_type='LANE', request_section_id=request_section_id,
                          record_count=len(excel_data) - 1,
                          created_by=user, rate_type=rate_type)
        file.save()
        RequestDataImportQueueManager.insert_all_lanes(excel_data, user, file)
        # validate_lanes(file)
        validate()

        return Response(status=status.HTTP_201_CREATED)


class RequestSectionLanePricingPointPreprocessingView(views.APIView):
    def post(self, request, *args, **kwargs):
        user = self.request.user
        rate_type = kwargs.get("rate_type")

        excel_file = request.FILES["import_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Pricing Points Template"]
        excel_data = list()

        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(cell.value)
            excel_data.append(row_data)
            file = ImportFile(file_name=excel_file.name, uni_type='PRICINGPOINT', request_section_id=0,
                              record_count=len(excel_data) - 1,
                              created_by=user, rate_type=rate_type)
        # Insert all Pricing Points
        # user = args[0]
        # file = args[1]
        weight_breaks = excel_data[0][10:]
        weight_break_json = {}
        header = None
        lanes = []

        pricingpoint_count = 0
        for idx, row in enumerate(excel_data):
            rates = row[10:]
            if idx == 0:
                uni_type = 'HEADER'
                weight_break_json = json.dumps(dict(zip(weight_breaks, rates)))
                uni_status = 'UNPROCESSED'
            else:
                if not all_elem_same_value(row[5:10], None):
                    uni_type = 'DATA'
                    uni_status = 'UNPROCESSED'
                    weight_break_json = json.dumps(dict(zip(weight_breaks, rates)))
                    pricingpoint_count += 1
                else:
                    uni_type = 'LANE'
                    uni_status = 'SKIPPED'

            lane = RequestSectionLanePricingPointImportQueue(request_section_id=row[0],
                                                             section_name=row[1],
                                                             request_section_lane_id=row[2],
                                                             origin_point_code=row[3],
                                                             destination_point_code=row[4],
                                                             request_section_lane_pricing_point_id=row[5],
                                                             origin_post_code_id=row[6],
                                                             origin_postal_code_name=row[7],
                                                             destination_post_code_id=row[8],
                                                             destination_postal_code_name=row[9],
                                                             weight_break=weight_break_json,
                                                             created_by=user,
                                                             uni_type=uni_type,
                                                             file=file,
                                                             uni_status=uni_status,
                                                             status_message={},
                                                             initial_rec_order=idx)
            if uni_type != "HEADER":
                lanes.append(lane)
            else:
                header = lane

        # Validate lanes
        for pricing_point in lanes:
            # Old Validatorion lanes
            if pricing_point.origin_postal_code_name and pricing_point.destination_postal_code_name:
                request_section_id = pricing_point.request_section_id
                request_section = RequestSection.objects.get(pk=request_section_id)
                sub_service_level_id = request_section.sub_service_level.sub_service_level_id
                lane = RequestSectionLane.objects.filter(
                    request_section_lane_id=pricing_point.request_section_lane_id).first()
                try:
                    do_validate_pricing_point(file, header, pricing_point, lane=lane)
                    resolve_pricing_point_ids(pricing_point)

                    # Service Matrix Validation
                    service_matrix_results = validate_new_lane(service_level_id=sub_service_level_id,
                                                               orig_point_type="Postal Code",
                                                               origin_point_value=pricing_point.origin_post_code_id,
                                                               dest_point_type="Postal Code",
                                                               destination_point_value=pricing_point.destination_post_code_id)
                    if service_matrix_results["FLAGGED"] == 1:
                        pricing_point.uni_status = "FLAGGED"
                    if service_matrix_results["UNSERVICEABLE"] == 1:
                        pricing_point.uni_status = "UNSERVICEABLE"
                except Exception as e:
                    pricing_point.uni_status = 'INVALID'
                    pricing_point.status_message = {"PP999": "General validation exception"}

        # serialize response
        results = []
        for pricing_point in lanes:
            results.append({
                'request_section_id': pricing_point.request_section_id,
                'section_name': pricing_point.section_name,
                'request_section_lane_id': pricing_point.request_section_lane_id,
                'origin_point_code': pricing_point.origin_point_code,
                'destination_point_code': pricing_point.destination_point_code,
                'request_section_lane_pricing_point_id': pricing_point.request_section_lane_pricing_point_id,
                'origin_post_code_id': pricing_point.origin_post_code_id,
                'origin_postal_code_name': pricing_point.origin_postal_code_name,
                'destination_post_code_id': pricing_point.destination_post_code_id,
                'destination_postal_code_name': pricing_point.destination_postal_code_name,
                'weight_break': pricing_point.weight_break,
                'uni_status': pricing_point.uni_status,
                'status_message': pricing_point.status_message
            })
        return Response(status=status.HTTP_200_OK, data={'lanes': results})


class RequestLanePricingPointImportView(views.APIView):
    @transaction.atomic()
    def post(self, request, *args, **kwargs):
        user = self.request.user
        rate_type = kwargs.get("rate_type")
        request_section_id = kwargs.get("request_section_id")
        excel_file = request.FILES["import_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Pricing Points Template"]
        excel_data = list()

        for row in worksheet.iter_rows():
            row_data = list()

            for cell in row:
                # cell_value = str(cell.value)

                # row_data.append(cell.value if cell.value is not None else None)
                row_data.append(cell.value)

            # excluding here the lanes from file, since we need them only user's convenience
            # if not all_elem_same_value(row_data[5:10], None):
            excel_data.append(row_data)
        # record_count is always -1 because we supposed to have HEADER row, which is not counting toward total number
        file = ImportFile(file_name=excel_file.name, uni_type='PRICINGPOINT', request_section_id=request_section_id,
                          record_count=len(excel_data) - 1,
                          created_by=user, rate_type=rate_type)
        file.save()
        file.record_count = RequestDataImportQueueManager.insert_all_pricing_points(excel_data, user, file)

        file.save()
        return Response(status=status.HTTP_201_CREATED)


class RequestLaneImportConfirmView(views.APIView):

    def get(self, request, *args, **kwargs):
        user = self.request.user
        request_section_id = kwargs.get("request_section_id")

        RequestDataImportQueueManager.commit_lane_import(request_section_id, user)

        return Response(status=status.HTTP_200_OK)


class ImportResults(views.APIView):

    def get(self, request, *args, **kwargs):

        request_section_id = kwargs.get("request_section_id")

        # getting most recent file in status VALID or INVALID, disregard to the uni_type (LANE or PRICINGPOINT)
        imported_file = ImportFile.objects.filter(request_section_id=request_section_id,
                                                  uni_status__in=['VALID', 'INVALID']).order_by('-created_on').first()

        if imported_file:
            file_uni_type = imported_file.uni_type

            if file_uni_type == 'LANE':
                return RequestDataImportQueueManager.get_lanes_import_results(request_section_id, imported_file)
            elif file_uni_type == 'PRICINGPOINT':
                return RequestDataImportQueueManager.get_pricingpoints_import_results(request_section_id, imported_file)
        # if there is no files - returning 404
        return Response({"Reason": "NO_RESULTS_FOUND"}, status=status.HTTP_404_NOT_FOUND)


class RequestLanePricingPointImportConfirmView(views.APIView):

    def get(self, request, *args, **kwargs):
        user = self.request.user
        request_section_id = kwargs.get("request_section_id")

        RequestDataImportQueueManager.commit_pricingpoint_import(request_section_id, user)

        return Response(status=status.HTTP_200_OK)


class RequestLaneImportCancelView(views.APIView):

    def get(self, request, *args, **kwargs):
        user = self.request.user
        request_section_id = kwargs.get("request_section_id")

        RequestDataImportQueueManager.cancel_import(request_section_id, user)

        return Response(status=status.HTTP_200_OK)


class RequestLaneImportStatusView(views.APIView):

    def get(self, request, *args, **kwargs):
        # user = self.request.user
        request_section_id = kwargs.get("request_section_id")
        # request_section = RequestSection.objects.get(pk=request_section_id)
        # sub_service_level_id = request_section.sub_service_level.sub_service_level_id
        # all_validation_results = {"FLAGGED": 0, "UNSERVICEABLE": 0}
        results = []
        for file in ImportFile.objects.filter(request_section_id=request_section_id):
            # status_array = list(
            #     RequestSectionLaneImportQueue.objects.filter(file=file).values(
            #         'uni_status', 'status_message'))
            status_array = list()
            if file.uni_type == 'LANE':
                status_array = list(
                    RequestSectionLaneImportQueue.objects.filter(file=file).exclude(uni_type='HEADER').values(
                        'uni_status').annotate(count=Count('uni_status')))
                # logging.info('status_array: %s \n', status_array)
            elif file.uni_type == 'PRICINGPOINT':
                status_array = list(
                    RequestSectionLanePricingPointImportQueue.objects.filter(file=file).exclude(
                        uni_type='HEADER').values(
                        'uni_status').annotate(
                        count=Count('uni_status')))

            status_array.append({"uni_status": "DIRECTIONAL", "count": file.directional_lane_count})
            status_array.append({"uni_status": "DUPLICATED", "count": file.duplicate_lane_count})
            status_array.append({"uni_status": "IS BETWEEN", "count": file.between_lane_count})
            status_array.append({"uni_status": "FLAGGED", "count": file.flagged_count})
            status_array.append({"uni_status": "UNSERVICEABLE", "count": file.unserviceable_count})

            results.append(
                {"file": {'id': file.id, 'name': file.file_name, 'status': file.uni_status, 'type': file.uni_type,
                          'results': status_array}})

        # status_array = list(
        #     RequestSectionLaneImportQueue.objects.filter(request_section_id=request_section_id).values(
        #         'uni_status').annotate(
        #         count=Count('uni_status')))
        return Response(results, status=status.HTTP_200_OK)


class RequestLaneImportDetailedStatusView(views.APIView):

    def get(self, request, *args, **kwargs):
        user = self.request.user
        status_array = list(
            RequestSectionLaneImportQueue.objects.filter(created_by=user).values('uni_status').annotate(
                count=Count('uni_status')))
        return Response(status_array, status=status.HTTP_200_OK)


class SearchRequestSectionLaneGroupsWithFilterPyodbcView(views.APIView):
    def get(self, request, *args, **kwargs):
        request_section_id = kwargs.get('request_section_id')
        group_type_name = kwargs.get('group_type_name')
        search_term = kwargs.get('search_term')

        filtered_items = []
        if request_section_id == '-1':
            filtered_items = list(LocationTreeView.objects.filter(point_type_name=group_type_name).values())
        else:
            request_section_id = int(request_section_id)
            service_level_code = RequestSection.objects.get(
                pk=request_section_id).sub_service_level.service_level.service_level_code
            base_restrictions = getattr(settings, 'SERVICE_LEVEL_BASE_POINT_RESTRICTION_FOR_' + service_level_code,
                                        None).split(',')
            filtered_items = list(LocationTreeView.objects.filter(point_type_name=group_type_name, country__in=base_restrictions).values())

        items_with_search_term = []
        for item in filtered_items:
            if search_term.lower() in item['name'].lower():
                items_with_search_term.append(item)

        return Response(items_with_search_term, status=status.HTTP_200_OK)


class SearchRequestSectionLanePointsWithFilterPyodbcView(views.APIView):
    def get(self, request, *args, **kwargs):
        filter_point_type_id = kwargs.get('filter_point_type_id')
        filter_point_id = kwargs.get('filter_point_id')
        response_point_type_id = kwargs.get('response_point_type_id')
        search_term = kwargs.get('search_term')

        point_type_mapping = {
            'Country': 'country_id',
            'Region': 'region_id',
            'Province': 'province_id',
            'Terminal': 'terminal_id',
            'Basing Point': 'basing_point_id',
            'Service Point': 'service_point_id',
            'Postal Code': 'postal_code_id',
            'Customer Zone': 'customer_zone_id',
            'Sub Postal Code': 'sub_postal_code_id'
        }
        filter_point_type_name = PointType.objects.get(pk=filter_point_type_id).point_type_name

        filter_condition = {
            'point_type_id': response_point_type_id,
            point_type_mapping[filter_point_type_name]: filter_point_id
        }

        filtered_items = list(LocationTreeView.objects.filter(**filter_condition).values())
        items_with_search_term = []
        for item in filtered_items:
            if search_term.lower() in item['name'].lower():
                items_with_search_term.append(item)

        return Response(items_with_search_term, status=status.HTTP_200_OK)


class SearchRequestSectionLanePointsPyodbcView(views.APIView):

    def _get_base_restrictions(self, service_level_id):
        service_level = ServiceLevel.objects.filter(service_level_id=service_level_id).first()
        service_level_code = service_level.service_level_code if service_level else None
        base_restriction_prefix = 'SERVICE_LEVEL_BASE_POINT_RESTRICTION_FOR_'
        country_codes = getattr(settings, base_restriction_prefix + service_level_code, None)
        return country_codes

    def get(self, request, *args, **kwargs):
        group_type = kwargs.get("group_type")
        group_id = kwargs.get("group_id")
        point_type = kwargs.get("point_type")
        point_name = kwargs.get("point_name")

        if group_id and group_type:  # already in canada
            service_level_id = request.query_params.get('service_level_id')
            # get the base restriction values cached from the azure key vault into settings
            countries = self._get_base_restrictions(service_level_id) if service_level_id else ''
            if countries == '1':
                group_type = "Country"
                group_id = 1

        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        query = queries.SEARCH_REQUEST_SECTION_LANE_POINTS.format(
            group_type, group_id, point_type, point_name)

        cursor.execute(query)
        raw_data = cursor.fetchone()
        payload = json.loads(raw_data[0]) if raw_data[0] else []

        return Response(payload, status=status.HTTP_200_OK)


class SearchOriginPostalCodesPyodbcView(views.APIView):

    def get(self, request, *args, **kwargs):
        request_section_lane_id = kwargs.get("request_section_lane_id")
        postal_code = kwargs.get("postal_code")
        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        query = queries.SEARCH_ORIGIN_POSTAL_CODE.format(
            request_section_lane_id, postal_code)

        cursor.execute(query)
        raw_data = cursor.fetchone()
        payload = json.loads(raw_data[0]) if raw_data[0] else []
        return Response(payload, status=status.HTTP_200_OK)


class SearchDestinationPostalCodesPyodbcView(views.APIView):

    def get(self, request, *args, **kwargs):
        request_section_lane_id = kwargs.get("request_section_lane_id")
        postal_code = kwargs.get("postal_code")
        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        query = queries.SEARCH_DESTINATION_POSTAL_CODE.format(
            request_section_lane_id, postal_code)

        cursor.execute(query)
        raw_data = cursor.fetchone()
        payload = json.loads(raw_data[0]) if raw_data[0] else []

        return Response(payload, status=status.HTTP_200_OK)


class GetRequestSectionLanePricingPointsPyodbcView(views.APIView):

    def get(self, request, *args, **kwargs):
        request_section_lane_id = kwargs.get("request_section_lane_id")
        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        query = queries.GET_PRICING_POINTS.format(request_section_lane_id)

        data = getFormattedRowResults(cursor, query)
        return Response(data, status=status.HTTP_200_OK)


class PathRequestSectionLanePricingPointCostOverridePyodbcView(views.APIView):

    def patch(self, request, *args, **kwargs):
        request_section_lane_pricing_point_ids = request.data["request_section_lane_pricing_point_ids"]

        request_data = request.data

        request_section_lane_pricing_point_ids = ','.join(map(str, request_section_lane_pricing_point_ids))
        cost_override_pickup_count = request_data.get("cost_override_pickup_count", 0)
        cost_override_delivery_count = request_data.get("cost_override_delivery_count")
        cost_override_dock_adjustment = request_data.get("cost_override_dock_adjustment", "[]")
        cost_override_margin = json.dumps(request_data.get("cost_override_margin", "[]"))
        cost_override_density = json.dumps(request_data.get("cost_override_density", "[]"))
        cost_override_pickup_cost = json.dumps(request_data.get("cost_override_pickup_cost", "[]"))
        cost_override_delivery_cost = json.dumps(request_data.get("cost_override_delivery_cost", "[]"))
        cost_override_accessorials_value = json.dumps(request_data.get("cost_override_accessorials_value", "[]"))
        cost_override_accessorials_percentage = json.dumps(
            request_data.get("cost_override_accessorials_percentage", "[]"))

        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        query = queries.UPDATE_PRICING_POINTS_COST_OVERRIDE.format(cost_override_accessorials_percentage,
                                                                   cost_override_accessorials_value,
                                                                   cost_override_delivery_cost,
                                                                   cost_override_delivery_count,
                                                                   cost_override_density,
                                                                   cost_override_dock_adjustment,
                                                                   cost_override_margin,
                                                                   cost_override_pickup_cost,
                                                                   cost_override_pickup_count,
                                                                   request_section_lane_pricing_point_ids)

        cursor.execute(query)
        cursor.commit()
        # raw_data = cursor.fetchone()
        # payload = json.loads(raw_data[0]) if raw_data[0] else []

        return Response(status=status.HTTP_200_OK)


class GetRequestSectionLanePricingPointsHistoryPyodbcView(views.APIView):

    def get(self, request, *args, **kwargs):
        request_section_lane_id = kwargs.get("request_section_lane_id")
        version_num = kwargs.get("version_num")
        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        query = queries.GET_PRICING_POINTS_HISTORY.format(
            request_section_lane_id, version_num)

        cursor.execute(query)
        raw_data = cursor.fetchone()
        payload = json.loads(raw_data[0]) if raw_data[0] else []

        ser = RequestSectionLanePricingPointSerializer(data=payload, many=True)

        if ser.is_valid():
            return Response(ser.validated_data, status=status.HTTP_200_OK)
        else:
            return Response({"status": "Failure", "error": ser.errors}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GetRequestSectionLaneChangeCountPyodbcView(views.APIView):

    def get(self, request, *args, **kwargs):
        request_section_id = kwargs.get("request_section_id")
        context_id = kwargs.get("context_id")
        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        query = queries.GET_REQUEST_SECTION_LANE_CHANGES_COUNT.format(
            request_section_id, context_id)

        cursor.execute(query)
        raw_data = cursor.fetchone()
        payload = json.loads(raw_data[0]) if raw_data[0] else []

        return Response(payload, status=status.HTTP_200_OK)


class GetRequestSectionLanePricingPointDestinationPyodbcView(views.APIView):

    def get(self, request, *args, **kwargs):
        destination_request_section_id = kwargs.get(
            "destination_request_section_id")
        request_section_lane_pricing_point_id = kwargs.get(
            "request_section_lane_pricing_point_id")
        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        query = queries.GET_PRICING_POINT_DESTINATION.format(
            request_section_lane_pricing_point_id, destination_request_section_id)

        cursor.execute(query)
        raw_data = cursor.fetchone()
        payload = json.loads(raw_data[0]) if raw_data[0] else []

        return Response(payload, status=status.HTTP_200_OK)


class SpeedsheetInformationUpdateView(views.APIView):

    @transaction.atomic
    def put(self, request, *args, **kwargs):
        request_id = kwargs.get("RequestID")
        request_information = request.data.get("request_information", {})
        is_macro_save = request.data.get("is_macro_save", False)

        if request_information:
            request_information_modified = False

            request_instance = Request.objects.select_related().filter(
                request_id=request_id).first()
            request_information_instance = request_instance.request_information

            new_language_code = request_information.get('language_code')
            new_service_level_code = request_information.get('service_level_code')
            new_speedsheet_name = request_information.get('speedsheet_name')

            # updating name
            if new_speedsheet_name:
                request_instance.speedsheet_name = new_speedsheet_name

            # updating language code if specified
            if new_language_code:
                language = Language.objects.filter(language_code=new_language_code).first()
                # request_instance.language = language
                request_information_instance.language = language
                request_information_modified = True
            # updating service level if specified
            if new_service_level_code:
                service_level = ServiceLevel.objects.filter(service_level_code=new_service_level_code).first()

                customer_instance = request_information_instance.customer
                customer_instance.service_level = service_level
                customer_instance.save()
                request_information_modified = True

            if request_information_modified:
                request_information_instance.save()
            # TODO to take a look on the object saving method, and make sure we are not doing unncessary heavy queries
            # TODO maybe reduce number of nested objects
            request_instance.save()

        return Response({"status": "Success"}, status=status.HTTP_200_OK)


class RequestViewSet(GetQuerySetMixin, GetSerializerClassMixin, RetrieveHistoryMixin, RevertVersionMixin,
                     viewsets.GenericViewSet, mixins.UpdateModelMixin):
    serializer_class = RequestSerializer
    serializer_class_history = RequestHistorySerializer
    queryset = Request.objects.filter(is_inactive_viewable=True)
    queryset_history = RequestHistory.objects.all()
    lookup_field = 'request_id'

    def get_queryset(self):
        return super().get_queryset()

    def get_serializer_class(self):
        return super().get_serializer_class()

    @action(methods=['get'], detail=True, url_path='history', url_name='history')
    def retrieve_history(self, request, *args, **kwargs):
        return super().retrieve_history(request, *args, **kwargs)

    @transaction.atomic
    @action(methods=['put'], detail=True, url_path='revert/(?P<version_num>[^/.]+)', url_name=r'revert')
    def revert_version(self, request, *args, **kwargs):
        request_instance = self.get_object()
        request_history_instance = self.queryset_history.filter(
            **kwargs).first()
        request_serializer = self.get_serializer_class()

        if not request_history_instance:
            return Response("History instance not found, check version number.", status=status.HTTP_404_NOT_FOUND)

        for field, serializer in {"request_profile": RequestProfileSerializer,
                                  "request_lane": RequestLaneRevertSerializer}.items():
            revert_instance(getattr(request_instance, field), getattr(
                request_history_instance, field + "_version"), serializer)

        request_serializer_data = revert_instance(
            request_instance, request_history_instance, request_serializer)

        request_filter_kwargs = {
            "request_number": request_instance.request_number, "is_latest_version": True}

        latest_request_history_instance = RequestHistory.objects.filter(
            **request_filter_kwargs).first()

        for field, model in {"request_profile_version": RequestProfileHistory}.items():
            setattr(latest_request_history_instance, field,
                    model.objects.filter(**request_filter_kwargs).first())

        latest_request_history_instance.save()

        request_id = kwargs.get("request_id")
        version_num = kwargs.get("version_num")

        with connection.cursor() as cursor:
            try:
                cursor.execute("EXEC [dbo].[RequestLane_Revert] %s, %s", [request_id, version_num])
                cursor.commit()
            finally:
                cursor.close()

        return Response(request_serializer_data, status=status.HTTP_200_OK)


class RequestReassignUpdateView(generics.GenericAPIView):
    serializer_class = AccountOwnerSerializer
    queryset = Request.objects.filter(is_inactive_viewable=True).select_related(
        'request_information__customer__account')
    lookup_field = 'request_id'

    def patch(self, request, *args, **kwargs):
        request_instance = self.get_object()
        account_instance = request_instance.request_information.customer.account

        if not account_instance:
            return Response(
                f"Account assosciated with customer_id {account_instance.request_information.customer_id} does not exist",
                status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(account_instance, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({"status": "Success"}, status=status.HTTP_200_OK)


class RequestStatusReassignUpdateView(views.APIView):
    def update(self, request, *args, **kwargs):
        user_id = self.request.user.user_id
        request_id = kwargs.get("request_id")
        # Not clear where this is called from the front-end, but just a straight pass-through to update Request.CurrentEditorID
        # TODO: rebuild a process where a request to reassign is created, then when approved make the reassignment
        # TODO: need to retrieve the user to reassign to; this just reassigns to the calling user
        conn = pyodbc_connection()
        reassign_str = f"""UPDATE dbo.Request SET CurrentEditorID = {user_id} WHERE RequestID = {request_id} """
        conn.execute(update_str)
        return Response({}, status=status.HTTP_200_OK)


class RequestStatusCurrentEditorUpdateView(views.APIView):
    # TODO: save the request to change assigned user here
    def post(self, request, *args, **kwargs):
        return Response({}, status=status.HTTP_200_OK)


class RequestStatusRequestEditorRightView(views.APIView):

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        request_number = request.data.get('request_number')
        request_id = request.data.get('request_id')
        user_name = self.request.user.user_name
        user_id = self.request.user.user_id
        # TODO: clearly was not working as written; request_number and id are not passed in as parameters

        notification_message = {
            "message": f"{user_name} has requested Editor Rights of RRF {request_number}. Please ensure that you have saved your work before clicking 'Approve' as any unsaved changes made will be lost.",
            "args": {
                "endpoint": f"requesteditor/update/{request_id}/",
                "request_id": request_id,
                "actions": [
                    {
                        "text": "Approve",
                        "payload": {
                            "action": "Approve",
                            "current_editor": user_instance.user_id
                        },
                        "alert": {
                            "alert_name": "requestApproved",
                            "alert_data": request_number
                        }
                    },
                    {
                        "text": "Decline",
                        "payload": {
                            "action": "Decline",
                            "current_editor": user_id
                        },
                        "alert": {
                            "alert_name": "requestDeclined",
                            "alert_data": request_number
                        }
                    }
                ]
            }
        }
        # TODO: lookup current editor id
        current_editor_id = 59
        request_instance = Request.objects.filter(request=request_id)
        notification_instance = Notification.objects.create(
            user=current_editor_id, message=json.dumps(notification_message))
        request_editor_right_instance = RequestEditorRight.objects.create(
            user=user_instance, request=request_instance, notification=notification_instance)

        return Response({"status": "Success"}, status=status.HTTP_201_CREATED)


class RequestSectionViewSet(viewsets.GenericViewSet, mixins.CreateModelMixin, mixins.UpdateModelMixin):
    serializer_class = RequestSectionSerializer
    queryset = RequestSection.objects.filter(is_inactive_viewable=True)
    lookup_field = 'request_section_id'


class RequestSectionCreateUpdateDuplicateView(views.APIView):
    @transaction.atomic
    def put(self, request, *args, **kwargs):
        try:
            request_sections = request.data.get("request_sections", [])
            clear_cost_param_array = []
            num_sections_changed = False

            for request_section in request_sections:
                # FE is using RateBase dropdown to show EffectiveDate Dropdown and sending back RateBase ID from EffectiveDate
                effective_date = request_section.get('effective_date', None)
                request_section['rate_base'] = effective_date

                request_section_id = request_section.get("request_section_id")
                if not request_section_id:
                    # PAC-1835
                    weight_break_header = WeightBreakHeader.objects.filter(
                        weight_break_header_id=request_section.get('weight_break_header', {})).first()
                    request_section['unit_factor'] = weight_break_header.unit_factor
                    request_section['as_rating'] = weight_break_header.as_rating
                    request_section['has_min'] = weight_break_header.has_min
                    request_section['has_max'] = weight_break_header.has_max
                    request_section['base_rate'] = weight_break_header.base_rate
                    request_section['weight_break_details'] = weight_break_header.levels
                    # PAC-1835 END
                    serializer = RequestSectionSerializer(data=request_section)
                    num_sections_changed = True
                else:
                    request_section_instance = RequestSection.objects.filter(
                        request_section_id=request_section_id).first()
                    if not request_section_instance:
                        return Response(f"RequestSection object with primary key '{request_section_id}' does not exist",
                                        status=status.HTTP_400_BAD_REQUEST)

                    key_field_changed = request_section.get(
                        "key_field_changed", False)
                    if key_field_changed:
                        clear_cost_param_array.append([request_section_id])

                    serializer = RequestSectionSerializer(
                        request_section_instance, data=request_section, partial=True)

                    if request_section.get("is_active") != request_section_instance.is_active:
                        num_sections_changed = True

                serializer.is_valid(raise_exception=True)
                request_section_instance = serializer.save()
                if num_sections_changed:
                    request_lane_instance = request_section_instance.request_lane
                    request_lane_instance.num_sections = RequestSection.objects.filter(
                        is_active=True, request_lane=request_section_instance.request_lane).count()
                    request_lane_instance.save()

            duplicate_request_sections = request.data.get(
                "duplicate_request_sections", [])

            request_id = request.data.get("request_id")
            raw_data = []

            if duplicate_request_sections and request_id:
                cnxn = pyodbc_connection()
                cursor = cnxn.cursor()
                duplicate_request_sections_param_array = [
                    [row["source_section_id"], row["destination_section_id"]] for row in duplicate_request_sections]
                cursor.execute("EXEC dbo.RequestSection_Copy ?, ?",
                               duplicate_request_sections_param_array, request_id)
                raw_data = cursor.fetchone()
                cursor.commit()

            if clear_cost_param_array and request_id:
                cnxn = pyodbc_connection()
                cursor = cnxn.cursor()
                cursor.execute("EXEC dbo.RequestSectionLane_Clear_Cost ?, ?",
                               clear_cost_param_array, request_id)
                cursor.commit()

            if not request_sections and not duplicate_request_sections:
                return Response({"status": "Unsuccesful"}, status=status.HTTP_400_BAD_REQUEST)

            payload = json.loads(raw_data[0]) if raw_data else []
            return Response(payload, status=status.HTTP_200_OK)

        except Exception as e:
            logging.warning("{} {}".format(type(e).__name__, e.args))
            return Response({"status": "Failure", "error": "{} {}".format(type(e).__name__, e.args)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class RequestSectionListView(generics.ListAPIView):
    serializer_class = RequestSectionRetrieveSerializer
    queryset = RequestSection.objects.filter(is_inactive_viewable=True)
    lookup_field = 'request_lane__request_number'
    lookup_url_kwarg = 'request_number'

    def get_queryset(self):
        return super().get_queryset().filter(**{self.lookup_field: self.kwargs[self.lookup_url_kwarg]})


class RequestSectionHistoryListVersionView(generics.ListAPIView):
    serializer_class = RequestSectionHistoryRetrieveSerializer
    queryset = RequestSectionHistory.objects.all().select_related('sub_service_level_version', 'rate_base_version',
                                                                  'override_class_version', 'equipment_type_version',
                                                                  'weight_break_header_version')
    lookup_field = 'request_id'

    def get_queryset(self):
        return super().get_queryset().filter(request__version__in=RequestHistory.objects.filter(**self.kwargs).values_list(
                'request_version_id', flat=True))


class StandardPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    page_size_query_description = 'The number of results to return per page.'
    max_page_size = 200


class RequestSectionLaneFilter(filters.FilterSet):
    request_section_id = filters.NumberFilter(
        field_name='request_section_id', lookup_expr='exact')
    origin_province_id = filters.NumberFilter(
        field_name='origin_province_id', lookup_expr='exact')
    origin_province_id__isnull = filters.BooleanFilter(
        field_name='origin_province_id', lookup_expr='isnull')
    origin_region_id = filters.NumberFilter(
        field_name='origin_region_id', lookup_expr='exact')
    origin_region_id__isnull = filters.BooleanFilter(
        field_name='origin_region_id', lookup_expr='isnull')
    origin_country_id = filters.NumberFilter(
        field_name='origin_country_id', lookup_expr='exact')
    origin_country_id__isnull = filters.BooleanFilter(
        field_name='origin_country_id', lookup_expr='isnull')
    origin_terminal_id = filters.NumberFilter(
        field_name='origin_terminal_id', lookup_expr='exact')
    origin_terminal_id__isnull = filters.BooleanFilter(
        field_name='origin_terminal_id', lookup_expr='isnull')
    origin_zone_id = filters.NumberFilter(
        field_name='origin_zone_id', lookup_expr='exact')
    origin_zone_id__isnull = filters.BooleanFilter(
        field_name='origin_zone_id', lookup_expr='isnull')
    origin_basing_point_id = filters.NumberFilter(
        field_name='origin_basing_point_id', lookup_expr='exact')
    origin_basing_point_id__isnull = filters.BooleanFilter(
        field_name='origin_basing_point_id', lookup_expr='isnull')
    origin_service_point_id = filters.NumberFilter(
        field_name='origin_service_point_id', lookup_expr='exact')
    origin_service_point_id__isnull = filters.BooleanFilter(
        field_name='origin_service_point_id', lookup_expr='isnull')
    origin_postal_code_id = filters.NumberFilter(
        field_name='origin_postal_code_id', lookup_expr='exact')
    origin_postal_code_id__isnull = filters.BooleanFilter(
        field_name='origin_postal_code_id', lookup_expr='isnull')
    origin_point_type_id = filters.NumberFilter(
        field_name='origin_point_type_id', lookup_expr='exact')
    destination_province_id = filters.NumberFilter(
        field_name='destination_province_id', lookup_expr='exact')
    destination_province_id__isnull = filters.BooleanFilter(
        field_name='destination_province_id', lookup_expr='isnull')
    destination_region_id = filters.NumberFilter(
        field_name='destination_region_id', lookup_expr='exact')
    destination_region_id__isnull = filters.BooleanFilter(
        field_name='destination_region_id', lookup_expr='isnull')
    destination_country_id = filters.NumberFilter(
        field_name='destination_country_id', lookup_expr='exact')
    destination_country_id__isnull = filters.BooleanFilter(
        field_name='destination_country_id', lookup_expr='isnull')
    destination_terminal_id = filters.NumberFilter(
        field_name='destination_terminal_id', lookup_expr='exact')
    destination_terminal_id__isnull = filters.BooleanFilter(
        field_name='destination_terminal_id', lookup_expr='isnull')
    destination_zone_id = filters.NumberFilter(
        field_name='destination_zone_id', lookup_expr='exact')
    destination_zone_id__isnull = filters.BooleanFilter(
        field_name='destination_zone_id', lookup_expr='isnull')
    destination_basing_point_id = filters.NumberFilter(
        field_name='destination_basing_point_id', lookup_expr='exact')
    destination_basing_point_id__isnull = filters.BooleanFilter(
        field_name='destination_basing_point_id', lookup_expr='isnull')
    destination_service_point_id = filters.NumberFilter(
        field_name='destination_service_point_id', lookup_expr='exact')
    destination_service_point_id__isnull = filters.BooleanFilter(
        field_name='destination_service_point_id', lookup_expr='isnull')
    destination_postal_code_id = filters.NumberFilter(
        field_name='destination_postal_code_id', lookup_expr='exact')
    destination_postal_code_id__isnull = filters.BooleanFilter(
        field_name='destination_postal_code_id', lookup_expr='isnull')
    destination_point_type_id = filters.NumberFilter(
        field_name='destination_point_type_id', lookup_expr='exact')
    is_published = filters.BooleanFilter(
        field_name='is_published', lookup_expr='exact')
    is_edited = filters.BooleanFilter(
        field_name='is_edited', lookup_expr='exact')
    is_duplicate = filters.BooleanFilter(
        field_name='is_duplicate', lookup_expr='exact')
    is_between = filters.BooleanFilter(
        field_name='is_between', lookup_expr='exact')
    do_not_meet_commitment = filters.BooleanFilter(
        field_name='do_not_meet_commitment', lookup_expr='exact')
    workflow_errors = filters.BooleanFilter(
        field_name='workflow_errors', lookup_expr='isnull', exclude=True)

    class Meta:
        model = RequestSectionLane
        fields = ['request_section_id', 'origin_province_id', 'origin_region_id', 'origin_country_id',
                  'origin_terminal_id', 'origin_zone_id', 'origin_basing_point_id', 'origin_service_point_id',
                  'origin_postal_code_id', 'origin_point_type_id', 'destination_province_id',
                  'destination_region_id', 'destination_country_id', 'destination_terminal_id', 'destination_zone_id',
                  'destination_basing_point_id', 'destination_service_point_id', 'destination_point_type_id',
                  'destination_postal_code_id', 'is_published', 'is_edited', 'is_duplicate', 'is_between',
                  'do_not_meet_commitment']


class RequestSectionLaneHistoryFilter(filters.FilterSet):
    request_section_id = filters.NumberFilter(
        field_name='request_section_version__request_section_id', lookup_expr='exact')
    origin_province_id = filters.NumberFilter(
        field_name='origin_province_version__province_id', lookup_expr='exact')
    origin_province_id__isnull = filters.BooleanFilter(
        field_name='origin_province_version_id', lookup_expr='isnull')
    origin_region_id = filters.NumberFilter(
        field_name='origin_region_version__region_id', lookup_expr='exact')
    origin_region_id__isnull = filters.BooleanFilter(
        field_name='origin_region_version_id', lookup_expr='isnull')
    origin_country_id = filters.NumberFilter(
        field_name='origin_country_version__country_id', lookup_expr='exact')
    origin_country_id__isnull = filters.BooleanFilter(
        field_name='origin_country_version_id', lookup_expr='isnull')
    origin_terminal_id = filters.NumberFilter(
        field_name='origin_terminal_version__terminal_id', lookup_expr='exact')
    origin_terminal_id__isnull = filters.BooleanFilter(
        field_name='origin_terminal_version_id', lookup_expr='isnull')
    origin_zone_id = filters.NumberFilter(
        field_name='origin_zone_version__zone_id', lookup_expr='exact')
    origin_zone_id__isnull = filters.BooleanFilter(
        field_name='origin_zone_version_id', lookup_expr='isnull')
    origin_basing_point_id = filters.NumberFilter(
        field_name='origin_basing_point_version__basing_point_id', lookup_expr='exact')
    origin_basing_point_id__isnull = filters.BooleanFilter(
        field_name='origin_basing_point_version_id', lookup_expr='isnull')
    origin_service_point_id = filters.NumberFilter(
        field_name='origin_service_point_version__service_point_id', lookup_expr='exact')
    origin_service_point_id__isnull = filters.BooleanFilter(
        field_name='origin_service_point_version_id', lookup_expr='isnull')
    origin_postal_code_id = filters.NumberFilter(
        field_name='origin_postal_code_version__postal_code_id', lookup_expr='exact')
    origin_postal_code_id__isnull = filters.BooleanFilter(
        field_name='origin_postal_code_version_id', lookup_expr='isnull')
    origin_point_type_id = filters.NumberFilter(
        field_name='origin_point_type_version__request_section_lane_point_type_id', lookup_expr='exact')
    destination_province_id = filters.NumberFilter(
        field_name='destination_province_version__province_id', lookup_expr='exact')
    destination_province_id__isnull = filters.BooleanFilter(
        field_name='destination_province_version_id', lookup_expr='isnull')
    destination_region_id = filters.NumberFilter(
        field_name='destination_region_version__region_id', lookup_expr='exact')
    destination_region_id__isnull = filters.BooleanFilter(
        field_name='destination_region_version_id', lookup_expr='isnull')
    destination_country_id = filters.NumberFilter(
        field_name='destination_country_version__country_id', lookup_expr='exact')
    destination_country_id__isnull = filters.BooleanFilter(
        field_name='destination_country_version_id', lookup_expr='isnull')
    destination_terminal_id = filters.NumberFilter(
        field_name='destination_terminal_version__terminal_id', lookup_expr='exact')
    destination_terminal_id__isnull = filters.BooleanFilter(
        field_name='destination_terminal_version_id', lookup_expr='isnull')
    destination_zone_id = filters.NumberFilter(
        field_name='destination_zone_version__zone_id', lookup_expr='exact')
    destination_zone_id__isnull = filters.BooleanFilter(
        field_name='destination_zone_version_id', lookup_expr='isnull')
    destination_basing_point_id = filters.NumberFilter(
        field_name='destination_basing_point_version_basing_point_id', lookup_expr='exact')
    destination_basing_point_id__isnull = filters.BooleanFilter(
        field_name='destination_basing_point_version_id', lookup_expr='isnull')
    destination_service_point_id = filters.NumberFilter(
        field_name='destination_service_point_version__service_point_id', lookup_expr='exact')
    destination_service_point_id__isnull = filters.BooleanFilter(
        field_name='destination_service_point_version_id', lookup_expr='isnull')
    destination_postal_code_id = filters.NumberFilter(
        field_name='destination_postal_code_version__postal_code_id', lookup_expr='exact')
    destination_postal_code_id__isnull = filters.BooleanFilter(
        field_name='destination_postal_code_version_id', lookup_expr='isnull')
    destination_point_type_id = filters.NumberFilter(
        field_name='destination_point_type_version__request_section_lane_point_type_id', lookup_expr='exact')
    is_published = filters.BooleanFilter(
        field_name='is_published', lookup_expr='exact')
    is_edited = filters.BooleanFilter(
        field_name='is_edited', lookup_expr='exact')
    is_duplicate = filters.BooleanFilter(
        field_name='is_duplicate', lookup_expr='exact')
    is_between = filters.BooleanFilter(
        field_name='is_between', lookup_expr='exact')
    do_not_meet_commitment = filters.BooleanFilter(
        field_name='do_not_meet_commitment', lookup_expr='exact')

    class Meta:
        model = RequestSectionLaneHistory
        fields = ['request_section_id', 'origin_province_id', 'origin_region_id', 'origin_country_id',
                  'origin_terminal_id', 'origin_zone_id', 'origin_basing_point_id', 'origin_service_point_id',
                  'origin_postal_code_id', 'origin_point_type_id', 'destination_province_id',
                  'destination_region_id', 'destination_country_id', 'destination_terminal_id', 'destination_zone_id',
                  'destination_basing_point_id', 'destination_service_point_id', 'destination_point_type_id',
                  'destination_postal_code_id', 'is_published', 'is_edited', 'is_duplicate', 'is_between',
                  'do_not_meet_commitment']


class RequestSectionLaneListView(generics.ListAPIView):
    serializer_class = RequestSectionLaneSerializer
    queryset = RequestSectionLane.objects.filter(
        is_inactive_viewable=True, is_active=True)
    pagination_class = StandardPagination
    filter_backends = [filters.DjangoFilterBackend, OrderingFilter]
    filterset_class = RequestSectionLaneFilter

    ordering_fields = ['origin_code', 'destination_code']

    # def get_queryset(self):
    #     return RequestSectionLane.objects.filter(
    #         is_inactive_viewable=True, is_active=True)

    def patch(self, request, *args, **kwargs):
        request_section_lane_id = kwargs.get("request_section_lane_id")
        action_type = kwargs.get("action_type")
        if action_type == 'exclude':
            RequestSectionLane.objects.filter(request_section_lane_id=request_section_lane_id).update(is_excluded=True)
        elif action_type == 'include':
            RequestSectionLane.objects.filter(request_section_lane_id=request_section_lane_id).update(is_excluded=False)

        return Response(status=status.HTTP_200_OK)


class RequestSectionLaneHistoryListVersionView(generics.ListAPIView):
    serializer_class = RequestSectionLaneHistoryRetrieveSerializer
    queryset = RequestSectionLaneHistory.objects.filter(is_inactive_viewable=True, is_active=True).select_related(
        'request_section_version',
        'origin_province_version', 'origin_region_version', 'origin_country_version', 'origin_terminal_version',
        'origin_zone_version', 'origin_basing_point_version', 'origin_service_point_version',
        'origin_postal_code_version', 'origin_point_type_version',
        'destination_province_version', 'destination_region_version', 'destination_country_version',
        'destination_terminal_version', 'destination_zone_version', 'destination_basing_point_version',
        'destination_service_point_version', 'destination_postal_code_version', 'destination_point_type_version')
    pagination_class = StandardPagination
    filter_backends = [filters.DjangoFilterBackend, OrderingFilter]
    filterset_class = RequestSectionLaneHistoryFilter

    ordering_fields = ['origin_code', 'destination_code']
    lookup_field = 'request_id'

    def get_queryset(self):
        return super().get_queryset().filter(request_section_version__in=RequestSectionHistory.objects.filter(
                request_version_id__in=RequestHistory.objects.filter(**self.kwargs).values_list(
                    'request_version_id', flat=True)))

def is_duplicate_lane(request_section_id, origin_type_id, origin_point_id, destination_type_id, destination_point_id):
    conn = pyodbc_connection()
    cursor = conn.cursor()
    is_duplicate_query = f"SELECT count(*) FROM dbo.RequestSectionLane WHERE OriginID='{origin_point_id}' AND OriginTypeID='{origin_type_id}' AND DestinationID = '{destination_point_id}' AND DestinationTypeID = '{destination_type_id}' AND IsActive = 1 And RequestSectionID = '{request_section_id}'"
    cursor.execute(is_duplicate_query)
    result = cursor.fetchone()
    return result[0] > 1



class CreateRequestSectionLanePyodbcView(views.APIView):

    def post(self, request, *args, **kwargs):
        lanes = request.data
        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()
        payload = []
        for data in lanes:
            try:
                request_section_id = data.get("request_section_id")
                orig_group_type_id = data.get("orig_group_type_id")
                orig_group_id = data.get("orig_group_id")
                orig_point_type_id = data.get("orig_point_type_id")
                orig_point_id = data.get("orig_point_id")
                dest_group_type_id = data.get("dest_group_type_id")
                dest_group_id = data.get("dest_group_id")
                dest_point_type_id = data.get("dest_point_type_id")
                dest_point_id = data.get("dest_point_id")
                is_between = data.get("is_between")

                # Validate new lane against service matrix
                request_section = RequestSection.objects.get(pk=request_section_id)
                sub_service_level_id = request_section.sub_service_level.sub_service_level_id

                origin_point_type_name = PointType.objects.filter(point_type_id=orig_point_type_id)[0].point_type_name
                destination_point_type_name = PointType.objects.filter(point_type_id=dest_point_type_id)[
                    0].point_type_name

                validation_results = validate_new_lane(service_level_id=sub_service_level_id,
                                                       orig_point_type=origin_point_type_name,
                                                       origin_point_value=orig_point_id,
                                                       dest_point_type=destination_point_type_name,
                                                       destination_point_value=dest_point_id)

                if validation_results["UNSERVICEABLE"] == 1:
                    return Response(
                        {'status': 'Failure', 'error': 'Lane(s) is unserviceable, and cannot be added to the request'},
                        status=status.HTTP_400_BAD_REQUEST)
                flagged = validation_results['FLAGGED']

                if not is_duplicate_lane(request_section_id=request_section_id,
                                         destination_type_id=dest_point_type_id,
                                         destination_point_id=dest_point_id,
                                         origin_type_id=orig_point_type_id,
                                         origin_point_id=orig_point_id):
                    cursor.execute("EXEC [dbo].[RequestSectionLane_Insert] ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?",
                                   request_section_id,
                                   orig_point_type_id,
                                   0,
                                   orig_point_id,
                                   dest_point_type_id,
                                   0,
                                   dest_point_id,
                                   is_between,
                                   flagged,
                                   0,
                                   0)
                    raw_data = cursor.fetchone()
                    # cursor.commit()
                    payload.append(raw_data[0])

            except Exception as e:
                cursor.rollback()
                logging.warning("{} {}".format(type(e).__name__, e.args))
                return Response({"status": "Failure", "error": "{} {}".format(type(e).__name__, e.args)},
                                status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        cursor.commit()
        return Response(payload, status=status.HTTP_200_OK)


class CreateRequestSectionLanePricingPointPyodbcView(views.APIView):

    def post(self, request, *args, **kwargs):

        try:
            data = request.data
            pricing_points = data.get("pricing_points")

            pricing_points_param_array = []

            if pricing_points:
                for row in pricing_points:
                    request_section_lane_id = row.get(
                        "request_section_lane_id")
                    origin_postal_code_id = row.get("origin_postal_code_id")
                    destination_postal_code_id = row.get(
                        "destination_postal_code_id")
                    request_section_id = RequestSectionLane.objects.get(pk=request_section_lane_id).request_section_id
                    request_section = RequestSection.objects.get(pk=request_section_id)
                    sub_service_level_id = request_section.sub_service_level.sub_service_level_id
                    service_matrix_validation_results = validate_new_lane(service_level_id=sub_service_level_id,
                                                                          orig_point_type="Postal Code",
                                                                          origin_point_value=origin_postal_code_id,
                                                                          dest_point_type="Postal Code",
                                                                          destination_point_value=destination_postal_code_id)

                    if service_matrix_validation_results["UNSERVICEABLE"] == 1:
                        return Response({'status': 'Failure',
                                         'errorMessage': 'Pricing Point(s) is unserviceable, and cannot be added to the request'},
                                        status=status.HTTP_400_BAD_REQUEST)
                    flagged = service_matrix_validation_results['FLAGGED']

                    pricing_points_param_array.append(
                        [request_section_lane_id, origin_postal_code_id, destination_postal_code_id, '{}', '{}',
                         flagged])

            if len(pricing_points_param_array) > 0:
                cnxn = pyodbc_connection()
                cursor = cnxn.cursor()
                for point in pricing_points_param_array:
                    table_type = [(point[0], point[1], point[2])]
                    params = (table_type, "{}", "{}", point[5])
                    cursor.execute(
                        "EXEC [dbo].[RequestSectionLanePricingPoint_Insert] @RequestSectionLanePricingPointTableType_Create=?, @UpdatedBy=?, @Comments=?, @Flagged=?",
                        params)
                cursor.commit()

            return Response({"status": "Success"}, status=status.HTTP_200_OK)
        except Exception as e:
            logging.warning("{} {}".format(type(e).__name__, e.args))
            return Response({"status": "Failure", "error": "{} {}".format(type(e).__name__, e.args)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UpdateRequestSectionLanePricingPointPyodbcView(views.APIView):

    def put(self, request, *args, **kwargs):

        try:
            data = request.data
            request_section_lane_id = data.get("request_section_lane_id")
            request_section_lane_pricing_points = data.get(
                "request_section_lane_pricing_points", '[]')
            destination_request_section_id = data.get(
                "destination_request_section_id")
            destination_request_section_lane_id = data.get(
                "destination_request_section_lane_id")
            is_move = data.get("is_move")

            is_active = data.get("is_active")

            cnxn = pyodbc_connection()
            cursor = cnxn.cursor()

            if destination_request_section_id is not None or destination_request_section_lane_id is not None:
                cursor.execute("EXEC [dbo].[RequestSectionLanePricingPoint_Copy] ?, ?, ?, ?, ?",
                               request_section_lane_pricing_points,
                               request_section_lane_id, destination_request_section_id, is_move,
                               destination_request_section_lane_id)
                cursor.commit()

            if is_active is not None and not is_active:
                cursor.execute("EXEC [dbo].[RequestSectionLanePricingPoint_Delete] ?, ?",
                               request_section_lane_id, request_section_lane_pricing_points)
                cursor.commit()

            return Response({"status": "Success"}, status=status.HTTP_200_OK)
        except Exception as e:
            logging.warning("{} {}".format(type(e).__name__, e.args))
            return Response({"status": "Failure", "error": "{} {}".format(type(e).__name__, e.args)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class WorkflowManagerView(views.APIView):

    @transaction.atomic
    def post(self, request, *args, **kwargs):

        try:
            save_comment(request, request.data.get("request_id", None))
        except Exception as error:
            return JsonResponse({"reason": str(error)}, status=500)

        current_request_status_type_id = request.data.get("current_request_status_type_id", 0)
        next_request_status_type_id = request.data.get("next_request_status_type_id")
        request_id = request.data.get("request_id", 0)
        pricing_engine_rating_filters = request.data.get("pricing_engine_rating_filters", None)

        request_instance = Request.objects.filter(
            request_id=request_id).select_related('request_information__customer__account',
                                                  'request_information__request_type').first()
        if not request_instance:
            return Response("Request not found.", status=status.HTTP_404_NOT_FOUND)

        request_status_instance = RequestStatus.objects.filter(request=request_instance).first()
        if not request_status_instance:
            return Response("RequestStatus not found.", status=status.HTTP_404_NOT_FOUND)

        current_request_status_type_instance = RequestStatusType.objects.filter(
            request_status_type_id=current_request_status_type_id).first()

        next_request_status_type_instance = RequestStatusType.objects.filter(
            request_status_type_id=next_request_status_type_id).first()

        workflow_manager = WorkflowManager(request_status=request_status_instance,
                                           request=request_instance,
                                           current_request_status_type=current_request_status_type_instance,
                                           next_request_status_type=next_request_status_type_instance,
                                           pricing_engine_rating_filters=pricing_engine_rating_filters)

        if current_request_status_type_instance:
            workflow_manager.close_request_queue()
        if next_request_status_type_instance:
            workflow_manager.generate_request_queue()

        secondary_pending_request_status_types = {
            'secondary_pending_drm': 'Pending DRM Approval',
            'secondary_pending_pcr': 'Pending PCR Approval',
            'secondary_pending_pc': 'Pending PC Approval',
            'secondary_pending_ept': 'Pending EPT Approval',
        }

        for secondary_pending_request_status_type_field, secondary_request_status_type_name in secondary_pending_request_status_types.items():
            is_secondary_status_requested = request.data.get(secondary_pending_request_status_type_field, 'false')
            # if request.data.get(secondary_pending_request_status_type_field) and request.data.get(secondary_pending_request_status_type_field) is True:
            if is_secondary_status_requested.lower() == 'true':
                secondary_next_request_status_type_instance = RequestStatusType.objects.filter(
                    request_status_type_name=secondary_request_status_type_name, is_secondary=True).first()
                if not secondary_next_request_status_type_instance:
                    return Response({"status": "Failure",
                                     "error": f"Secondary RequestStatusType with name {secondary_request_status_type_name} not found."},
                                    status=status.HTTP_404_NOT_FOUND)
                workflow_manager = WorkflowManager(
                    request_status=request_status_instance, request=request_instance,
                    next_request_status_type=secondary_next_request_status_type_instance)
                workflow_manager.generate_request_queue()

        return Response({"status": "Success"}, status=status.HTTP_200_OK)


class RequestCustomerAccountSynchronizeView(views.APIView):

    @transaction.atomic
    def put(self, request, *args, **kwargs):
        account_number = kwargs.get("account_number")
        request_number = kwargs.get("request_number")

        customer_instance = getattr(RequestInformation.objects.filter(
            request_number=request_number).first(), "customer", None)
        account_instance = Account.objects.filter(
            account_number=account_number).first()

        if not customer_instance:
            return Response({"status": "Failure", "error": "Customer not found."}, status=status.HTTP_404_NOT_FOUND)
        if not account_instance:
            return Response({"status": "Failure", "error": "Account not found."}, status=status.HTTP_404_NOT_FOUND)

        if Customer.objects.filter(service_level=customer_instance.service_level, account=account_instance).exists():
            return Response(
                {"status": "Failure", "error": "Customer assosciated with Account and ServiceLevel already exists."},
                status=status.HTTP_400_BAD_REQUEST)

        customer_account_mapping = {"city": "city", "customer_name": "account_name", "customer_alias": "account_alias",
                                    "customer_address_line_1": "address_line_1",
                                    "customer_address_line_2": "address_line_2", "postal_code": "postal_code"}

        for customer_field, account_field in customer_account_mapping.items():
            setattr(customer_instance, customer_field,
                    getattr(account_instance, account_field))

        customer_instance.account = account_instance
        customer_instance.save()

        return Response({"status": "Success"}, status=status.HTTP_200_OK)


def weight_breaks_to_json_array(weight_breaks):
    weight_break_list = []
    for weight_break in weight_breaks:
        # weight_break_list.append(f'{weight_break}:CCWT {weight_break}:true')
        weight_break_list.append(weight_break)
    json_array = json.dumps(weight_break_list)
    return json_array


class RequestDataImportQueueManager:

    @staticmethod
    @transaction.atomic()
    def insert_all_lanes(row_list, *args, **kwargs):
        user = args[0]
        file = args[1]
        weight_breaks = row_list[0][12:]
        logging.info('row_list: %s \n', row_list)
        logging.info('weight_breaks: %s \n', weight_breaks)
        for idx, row in enumerate(row_list):
            rates = row[12:]
            for r_idx, rate in enumerate(rates):
                try:
                    if rate:
                        rates[r_idx] = float(rate)
                    else:
                        rates[r_idx] = float(0)
                except ValueError:  # EK: This should be validated, or -1 inserted to trigger validation in async job
                    pass

            # EK: Refactor to bulk_insert
            # weight_break_json_array = weight_breaks_to_json_array(weight_breaks)
            weight_break_json_array = dict(zip(weight_breaks, rates))

            RequestSectionLaneImportQueue(request_section_id=row[0],
                                          section_name=row[1],
                                          request_section_lane_id=row[2],
                                          origin_group_type_name=row[3],
                                          origin_group_code=row[4],
                                          origin_point_type_name=row[5],
                                          origin_point_code=row[6],
                                          # origin_point_type_id=1,
                                          # origin_point_id=1,
                                          destination_group_type_name=row[7],
                                          destination_group_code=row[8],
                                          destination_point_type_name=row[9],
                                          destination_point_code=row[10],
                                          is_between=row[11],
                                          weight_break=json.dumps(weight_break_json_array),
                                          created_by=user,
                                          uni_type='HEADER' if idx == 0 else 'DATA',
                                          file=file).save()

    @staticmethod
    def insert_all_pricing_points(row_list, *args, **kwargs):
        user = args[0]
        file = args[1]
        weight_breaks = row_list[0][10:]
        weight_break_json = {}

        pricingpoint_count = 0
        for idx, row in enumerate(row_list):
            rates = row[10:]

            if idx == 0:
                uni_type = 'HEADER'
                weight_break_json = json.dumps(dict(zip(weight_breaks, rates)))
                uni_status = 'UNPROCESSED'
            else:
                if not all_elem_same_value(row[5:10], None):
                    uni_type = 'DATA'
                    uni_status = 'UNPROCESSED'
                    weight_break_json = json.dumps(dict(zip(weight_breaks, rates)))
                    pricingpoint_count += 1
                else:
                    uni_type = 'LANE'
                    uni_status = 'SKIPPED'

            RequestSectionLanePricingPointImportQueue(request_section_id=row[0],
                                                      section_name=row[1],
                                                      request_section_lane_id=row[2],
                                                      origin_point_code=row[3],
                                                      destination_point_code=row[4],
                                                      request_section_lane_pricing_point_id=row[5],
                                                      origin_post_code_id=row[6],
                                                      origin_postal_code_name=row[7],
                                                      destination_post_code_id=row[8],
                                                      destination_postal_code_name=row[9],
                                                      weight_break=weight_break_json,
                                                      created_by=user,
                                                      uni_type=uni_type,
                                                      file=file,
                                                      uni_status=uni_status,
                                                      status_message={},
                                                      initial_rec_order=idx).save()
        return pricingpoint_count

    @staticmethod
    @transaction.atomic()
    def commit_lane_import(request_section_id, *args, **kwargs):
        logging.info('commit_lane_import()\n')
        for file in ImportFile.objects.filter(request_section_id=request_section_id, uni_status='VALID'):
            # RequestSectionLaneImportQueue.objects.filter(file=file, uni_status__in=['VALID']).update(
            #     uni_status='IMPORTED')

            cnxn = pyodbc_connection()
            cursor = cnxn.cursor()
            updated_lanes = []

            # TODO replace constant UUID with new generated, no issues with this approach, but just not to confuse anybody
            context_id = 'ce3d7ec0e32449dabe246674f03a33ca'
            rate_table = file.rate_type.replace('-', '_')

            for request_section_lane in RequestSectionLaneImportQueue.objects.filter(file=file, uni_type='DATA').filter(
                    uni_status__in=['VALID', 'INVALID']):
                if request_section_lane.request_section_lane_id is None:
                    logging.info("request_section_lane_insert()\n")
                    is_flagged = file.flagged_count > 0
                    request_section_lane_insert(request_section_lane, rate_table=rate_table, flagged=is_flagged)
                else:
                    logging.info("else \n")

                    lane_status = 'None'
                    is_between = None
                    micro_save = None
                    macro_save = None

                    request_section_id = request_section_lane.request_section_id
                    request_section_lanes = '[{}]'.format(request_section_lane.request_section_lane_id)
                    weight_break_lower_bound = '[2000]'

                    updated_lanes.append(request_section_lane.request_section_lane_id)
                    # cursor.commit()
                    #
                    # cursor.execute("EXEC [dbo].[RequestSectionLane_Rate_Update] ?, ?, ?, ?",
                    #                request_section_lane.request_section_id,
                    #                request_section_lane.request_section_lane_id, None,
                    #                request_section_lane.weight_break)

                    print('committing rates for lane: {}'.format(request_section_lane.request_section_lane_id))

            file.uni_status = 'IMPORTED'
            file.updated_on = datetime.now(tz=timezone.utc)
            # cursor.execute("EXEC [dbo].[RequestSectionLane_History_Update_Proxy]")
            print(updated_lanes)
            file.save()

            NotificationManager.send_notification(file.created_by,
                                                  f"{file.record_count} lanes imported from {file.file_name}.",
                                                  {"type": "IMPORT_COMPLETED",
                                                   "request_section_id": file.request_section_id})

    @staticmethod
    @transaction.atomic()
    def commit_pricingpoint_import(request_section_id, *args, **kwargs):

        for file in ImportFile.objects.filter(request_section_id=request_section_id, uni_status='VALID'):
            new_pricing_points_param_array = []
            cnxn = pyodbc_connection()
            cursor = cnxn.cursor()
            updated_pricing_point = []
            context_id = 'ce3d7ec0e32449dabe246674f03a33ca'
            orig_type = 'None'
            orig_id = 0
            dest_type = 'None'
            dest_id = 0
            lane_status = 'None'
            request_section_lanes = '[]'
            operation = '='
            multiplier = 0
            weight_break_lower_bound = '[]'
            rate_table = file.rate_type.replace('-', '_')
            # rate_table = 'dr_rate'
            request_section_lane_pricing_points = []
            request_section_lane_pricing_point_kv = {}
            for pricing_point in RequestSectionLanePricingPointImportQueue.objects \
                    .filter(file=file, uni_type='DATA', uni_status__in=['VALID']):
                dr_rate = pricing_point.weight_break
                dr_fak = pricing_point.weight_break
                if pricing_point.request_section_lane_pricing_point_id is None:
                    new_pricing_points_param_array.append(
                        [int(pricing_point.request_section_lane_id),
                         pricing_point.origin_postal_code_id,
                         pricing_point.destination_postal_code_id,
                         dr_rate,
                         dr_fak])
                else:
                    lane_status = 'None'
                    request_section_id = pricing_point.request_section_id
                    request_section_lane_pricing_points = '[{}]'.format(
                        pricing_point.request_section_lane_pricing_point_id)
                    request_section_lane_pricing_point_kv[
                        pricing_point.request_section_lane_pricing_point_id] = pricing_point.weight_break

                    updated_pricing_point.append(pricing_point.request_section_lane_id)

            cnxn = pyodbc_connection()
            cursor = cnxn.cursor()
            micro_save = None
            macro_save = None
            cursor.execute(
                "EXEC [dbo].[RequestSectionLanePricingPoint_Staging_Update_Ex] ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?",
                request_section_id,
                context_id,
                orig_type,
                orig_id,
                dest_type,
                dest_id,
                lane_status,
                request_section_lanes,
                json.dumps(request_section_lane_pricing_point_kv),
                operation,
                multiplier,
                rate_table,
                weight_break_lower_bound,
                micro_save,
                macro_save
            )
            lane_status = 'None'
            is_between = None

            if len(new_pricing_points_param_array) > 0:
                request_section_lane_pricingpiont_insert(new_pricing_points_param_array)

            if len(updated_pricing_point) > 0:
                operation = None
            request_section_lane_pricing_points = '[]'
            multiplier = None
            rate_table = None
            cursor.execute(
                "EXEC [dbo].[RequestSectionLane_Staging_Update] ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?",
                request_section_id,
                context_id,
                orig_type,
                orig_id,
                dest_type,
                dest_id,
                lane_status,
                request_section_lanes,
                is_between,
                operation,
                multiplier,
                rate_table,
                weight_break_lower_bound,
                None,
                True
            )

            cursor.commit()
            print(updated_pricing_point)

            file.uni_status = 'IMPORTED'
            file.updated_on = datetime.now(tz=timezone.utc)
            file.save()
            NotificationManager.send_notification(file.created_by,
                                                  f"{file.record_count} pricing points imported from {file.file_name}.",
                                                  {"type": "IMPORT_COMPLETED",
                                                   "request_section_id": file.request_section_id})

    @staticmethod
    @transaction.atomic()
    def cancel_import(request_section_id, *args, **kwargs):
        for file in ImportFile.objects.filter(request_section_id=request_section_id,
                                              uni_status__in=['VALID', 'INVALID', 'UNPROCESSED']):
            # RequestSectionLaneImportQueue.objects.filter(file=file, uni_status__in=['VALID',]).update(
            #     uni_status='CANCELLED')
            file.uni_status = 'CANCELLED'
            file.save()
            NotificationManager.send_notification(
                file.created_by, f"Import of {file.record_count} lanes from {file.file_name} cancelled",
                {"type": "IMPORT_CANCELLED", "request_section_id": file.request_section_id})

    @staticmethod
    def get_lanes_import_results(request_section_id, file):

        output_file = io.BytesIO()
        invalid_header = False
        header_status_message = ''
        row_num = 0
        col_num = 0
        workbook = Workbook(output_file, {'in_memory': True})
        worksheet = workbook.add_worksheet(name='Lanes Template')
        request_section = RequestSection.objects.filter(request_section_id=request_section_id).values_list(
            "weight_break", "section_name", flat=False).get()
        weight_break = request_section[0]

        field_names = [
            'request_section_id', 'section_name', 'request_section_lane_id', 'origin_group_type_name',
            'origin_group_code',
            'origin_point_type_name', 'origin_point_code', 'destination_group_type_name', 'destination_group_code',
            'destination_point_type_name', 'destination_point_code', 'is_between']
        col_titles = [
            'RequestSectionID', 'SectionName', 'RequestSectionLaneID', 'OriginGroupTypeName', 'OriginGroupCode',
            'OriginPointTypeName', 'OriginPointCode', 'DestinationGroupTypeName', 'DestinationGroupCode',
            'DestinationPointTypeName', 'DestinationPointCode', 'IsBetween']

        for weight_break in json.loads(weight_break):
            split_weight_break = [x.split(':') for x in weight_break.split(' ')][0][0]

            col_titles.append(split_weight_break)
        col_titles.append('Status')
        col_titles.append('StatusMessage')

        for col_title in col_titles:
            worksheet.write(row_num, col_num, col_title)
            col_num += 1
        header_result = RequestSectionLaneImportQueue.objects.filter(file=file, uni_type='HEADER').first()
        if header_result.uni_status == 'INVALID':
            invalid_header = True
            header_status_message = header_result.status_message

        for request_section_lane in RequestSectionLaneImportQueue.objects.filter(file=file, uni_type='DATA'):

            row_num += 1
            col_num = 0

            for field_name in field_names:
                value = getattr(request_section_lane, field_name, None)
                worksheet.write(row_num, col_num, value)
                col_num += 1
            col_num = 12

            for key, value in json.loads(request_section_lane.weight_break).items():
                worksheet.write(row_num, col_num, value)
                col_num += 1

            worksheet.write(row_num, col_num, request_section_lane.uni_status)
            col_num += 1

            status_messages = ', '.join('[{0}]'.format(message) for message in
                                        list(json.loads(
                                            header_status_message if invalid_header else request_section_lane.status_message).values()))
            worksheet.write(row_num, col_num, status_messages)

        workbook.close()

        output_file.seek(0)

        response = HttpResponse(output_file.read(),
                                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = "attachment; filename=lanes_import_results.xlsx"

        output_file.close()

        file.uni_status = 'COMPLETED'
        file.updated_on = datetime.now(tz=timezone.utc)
        file.save()

        return response

    @staticmethod
    def get_pricingpoints_import_results(request_section_id, file):

        output_file = io.BytesIO()
        row_num = 0
        col_num = 0
        workbook = Workbook(output_file, {'in_memory': True})
        worksheet = workbook.add_worksheet(name='Pricing Points Template')
        request_section = RequestSection.objects.filter(request_section_id=request_section_id).values_list(
            "weight_break", "section_name", flat=False).get()
        header_result = RequestSectionLaneImportQueue.objects.filter(file=file, uni_type='HEADER').first()
        # weight_break = request_section[0]
        weight_break = header_result.weight_break

        file_export_header = RequestSectionLanePricingPointImportQueue.objects.filter(file=file,
                                                                                      uni_type__in=['HEADER']).first()

        field_names = [
            'request_section_id', 'section_name', 'request_section_lane_id', 'origin_point_code',
            'destination_point_code',
            'request_section_lane_pricing_point_id', 'origin_post_code_id', 'origin_postal_code_name',
            'destination_post_code_id', 'destination_postal_code_name']

        col_titles = [
            'RequestSectionID', 'SectionName', 'RequestSectionLaneID', 'OriginPointCode', 'DestinationPointCode',
            'RequestSectionLanePricingPointID', 'OriginPostCodeID', 'OriginPostalCodeName', 'DestinationPostCodeID',
            'DestinationPostalCodeName']

        col_titles_from_file = []
        for field_name in field_names:
            col_titles_from_file.append(getattr(file_export_header, field_name, None))

        for weight_break in json.loads(weight_break):
            split_weight_break = [x.split(':') for x in weight_break.split(' ')][0][0]

            col_titles_from_file.append(split_weight_break)
        col_titles_from_file.append('Status')
        col_titles_from_file.append('StatusMessage')

        for col_title in col_titles_from_file:
            worksheet.write(row_num, col_num, col_title)
            col_num += 1

        overriding_status_message = None
        if file_export_header and file_export_header.uni_status == 'INVALID':
            overriding_status_message = json.loads(file_export_header.status_message)

        for request_section_lane in RequestSectionLanePricingPointImportQueue \
                .objects.filter(file=file, uni_type__in=['LANE', 'DATA']).order_by('initial_rec_order'):

            row_num += 1
            col_num = 0
            for field_name in field_names:
                value = getattr(request_section_lane, field_name, None)
                worksheet.write(row_num, col_num, value)
                col_num += 1
            col_num = 10

            for key, value in json.loads(request_section_lane.weight_break).items():
                if request_section_lane.uni_type == 'DATA':
                    worksheet.write(row_num, col_num, value)
                col_num += 1
            if request_section_lane.uni_type == 'DATA':
                worksheet.write(row_num, col_num, request_section_lane.uni_status)

            col_num += 1

            status_messages = '[{0}]'.format(
                next(iter(overriding_status_message.items()))[1]) if overriding_status_message else ', '.join(
                '[{0}]'.format(message) for message in list(json.loads(request_section_lane.status_message).values()))
            worksheet.write(row_num, col_num, status_messages)

        workbook.close()

        output_file.seek(0)

        response = HttpResponse(output_file.read(),
                                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = "attachment; filename=pricing_points_import_results.xlsx"

        output_file.close()
        # TO DO Uncomment before commit
        file.uni_status = 'COMPLETED'
        file.updated_on = datetime.now(tz=timezone.utc)
        file.save()

        return response


class InitiateRRFView(views.APIView):
    def get(self, request, *args, **kwargs):
        def _run_query(query):
            cursor.execute(query)
            columns = [column[0] for column in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

        user_id = self.request.user.user_id
        cnxn = pyodbc_connection()
        cursor = cnxn.cursor()

        SERVICE_LEVELS_QUERY = """ select ServiceLevelID, ServiceLevelName, ServiceLevelCode from ServiceLevel; """

        payload = dict(
            tariff_types=["Non Customer Tariff", "Customer Tariff", "Tender"],
            status=["New", "Updated"],
            service_levels=_run_query(SERVICE_LEVELS_QUERY)
        )

        return Response(payload, status=status.HTTP_200_OK)


def __init__(self):
    pass
